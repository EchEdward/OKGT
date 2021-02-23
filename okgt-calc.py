import numpy as np

from scipy.integrate import simps
from scipy.constants import epsilon_0

import scipy.sparse as sparse
import scipy.sparse.linalg as linalg

from openpyxl import load_workbook

from initial_data import okgt_info, vl_info

import sys



def load_data():
    Katal = load_workbook(filename = 'Katal.xlsx')
    Kat = Katal['Kat']

    supports = {}
    conductors = {}

    s = (2,3,4,5,6,9,10,11,12,13,14,14)
    c = (18,19)
    check = lambda arr:sum([0 if Kat.cell(row=i,column=j).value is not None else 1 for i in arr])

    j = 2
    while  Kat.cell(row=1,column=j).value is not None or\
        Kat.cell(row=17,column=j).value is not None:
        if not check(s):
            supports[Kat.cell(row=1,column=j).value] = {
                "X":{
                    "A":Kat.cell(row=2,column=j).value,
                    "B":Kat.cell(row=3,column=j).value,
                    "C":Kat.cell(row=4,column=j).value,
                    "T1":Kat.cell(row=5,column=j).value,
                    "T2":Kat.cell(row=6,column=j).value,
                },
                "Y":{
                    "A":Kat.cell(row=9,column=j).value,
                    "B":Kat.cell(row=10,column=j).value,
                    "C":Kat.cell(row=11,column=j).value,
                    "T1":Kat.cell(row=12,column=j).value,
                    "T2":Kat.cell(row=13,column=j).value,
                },
                "type":Kat.cell(row=14,column=j).value,
                "isulator":Kat.cell(row=15,column=j).value,
            }
        if not check(c):
            conductors[Kat.cell(row=17,column=j).value] = {
                "r":Kat.cell(row=18,column=j).value,
                "R0":Kat.cell(row=19,column=j).value,
                "Id":Kat.cell(row=20,column=j).value,
            }
        j+=1

    return supports, conductors


k_supports, k_conductors = load_data()
#print(k_supports)


""" 
# graf of power station
Aps = sparse.lil_matrix((10,5),dtype=np.float64)
for i in range(5):
   Aps[i,i] = 1
   Aps[i+5,i+5] = -1

# graf of okgt with transmision line
Aokgt1n = sparse.lil_matrix((6,6),dtype=np.float64)
for i in range(6):
   Aokgt1n[i,i] = 1
Aokgt1k = - Aokgt1n

# graf of single conductive okgt 
Aokgt2n = sparse.lil_matrix((2,2),dtype=np.float64)
for i in range(6):
   Aokgt2n[i,i] = 1
Aokgt2k = - Aokgt2n

# graf of single dielectric okgt 
Aokgt3n = sparse.lil_matrix((1,1),dtype=np.float64)
for i in range(6):
   Aokgt3n[i,i] = 1
Aokgt3k = - Aokgt3n

# graf of transmision line without okgt 
Atrlnn = sparse.lil_matrix((5,5),dtype=np.float64)
for i in range(5):
   Atrlnn[i,i] = 1
Atrlnk = - Atrlnn 
"""

def get_vl_sector(key):
    lst = {}
    for name, value in vl_info.items():
        branches = {}
        for k, branch in value["branches"].items():
            branches[k] = (branch["supportN"],branch["supportK"])
        
        lst[name] = []

        for sector in value["sectors"]:
            if sector["type"] == "with_okgt":
                if sector["link_okgt"] == key:
                    d_sector = {k:v for k,v in sector.items()}
                    d_sector["branch_supports"] = branches[d_sector["link_branch"]]
                    d_sector["name_vl"] = name
                    lst[name].append(d_sector)

        if len(lst[name])==0:
            del lst[name]
    return lst   


def find_entry(dct, length):
    dl = 0.0001
    orders = {}
    for name, lst in dct.items():
        
        ln, lk = 0,  length
        order = []
        len_lst  = len(lst)
        for _ in range(len_lst):
            for i, val in enumerate(lst):
                if i in order: 
                    continue
                if ln-dl<=val['lengthN']<=ln+dl and val['lengthK']<=lk+dl:
                    order.append(i)
                    ln = val['lengthK']

            if len_lst == len(order):
                break
        else:
            
            raise Exception("Entries were not found")
            
        orders[name] = order

    return orders

def sector_entry(lst,branch):
    new_lst = [i for i,val in enumerate(lst) if val["link_branch"]==branch]
    if len(new_lst) == 0:
        raise Exception("Entries were not found 3")
    count = 0
    for i in new_lst:
        if lst[i]["supportN"]<lst[i]["supportK"]:
            count+=1
    #print(count, len(new_lst))
    if count != len(new_lst) and count != 0:
        raise Exception("Entries were not found 1")

    derection = count == len(new_lst)
    order = sorted(new_lst, key=lambda x: lst[x]["supportN"], reverse=not derection)
    for i in range(1,len(new_lst)):
        if lst[order[i]]["supportN"] != lst[order[i-1]]["supportK"]:
            raise Exception("Entries were not found 2")

    return order


def R_cc(r0,lg,dg,h):
    R = r0/(2*np.pi*lg*10**3)*(np.log((2*10**6*lg)/dg)+np.log(lg*10**3/2/h))
    return R



#print(R_cc(30,2.5,12,0.5),"aaaa")


pz = 30
carson_cashe = {}
mu0=np.pi*4*10**(-7)
jp=1j*2*np.pi*50
sh_int=np.arange(0, 1, 0.00001)
mp=(mu0*jp*(1/pz))**0.5
ro_fe =  0.1 #Ом*мм2/м
R_isol = 10**8

#print(sector_entry(vl_info["VL #4"]["sectors"],("3","4")))

#print(get_vl_sector("one"))
# Construct graph of okgt's links
nodes_position = {}
i, j = 0, 0
lst_conections = []
lst_cc_conections = []

lst_zy = []

to_okgt = {}
okgt_to_cc ={}

for (n, k), sectors in okgt_info.items():
    if n not in nodes_position:
        nodes_position[n] = (i,j)

    
    for ind, sector in enumerate(sectors):
        if sector["type"] == "VL":
            vl_sectors = get_vl_sector(sector["name"])

            ms = []

            dd = find_entry(vl_sectors,sector["length"])
            
            for name_v, val in dd.items():
                sm = 0
                for ind in val:
                    sm+= abs(vl_sectors[name_v][ind]['supportN']-vl_sectors[name_v][ind]['supportK'])
            
                ms.append(sm)

            print(dd)
            for 

            sys.exit()
            
            if len(set(ms))>1:
                nn = sector["name"]
                raise Exception(f"Double chain vls have diffrent length on sector {nn}")

            idd = {i:name for i, name in enumerate(vl_sectors.keys())}

            
            vl_sectors = [vl_sectors[idd[0]][i] for i in dd[idd[0]]]

            
            for m, vl_sector in enumerate(vl_sectors):
                N, K = vl_sector['supportN'], vl_sector['supportK']
                drct = int((K-N)/abs(K-N))
                gen = zip(range(abs(K-N)),range(N,K,drct),range(N+drct,K+drct,drct))
                dl = abs(vl_sector['lengthK']-vl_sector['lengthN'])/abs(K-N)
                
                for t, nc, kc in gen:
                    if ind == 0 and m==0 and t==0 and n in nodes_position:
                        i_old = nodes_position[n][0]
                        start = (i_old,j,-1)
                    else:
                        start = (i,j,-1)
                    i += 1
                    end = (i,j,1)
                    j += 1

                    to_okgt[(vl_sector["name_vl"],vl_sector["link_branch"],nc,kc)] = len(lst_conections)

                    d_row = {
                        "start": start,
                        "end": end,
                        "type":"VL",
                        'supportN': nc,
                        'supportK': kc,
                        "length":dl,
                        "link_branch":vl_sector["link_branch"],
                        "name_vl":vl_sector["name_vl"],
                        "link_okgt": sector["name"]
                        }
                    lst_conections.append(d_row)

  
        elif sector["type"] == "single_dielectric":
            if ind == 0 and n in nodes_position:
                i_old = nodes_position[n][0]
                start = (i_old,j,-1)
            else:
                start = (i,j,-1)
            i += 1
            end = (i,j,1)
            j += 1
            d_row = {"start": start, "end": end, "length":sector["length"], "type":sector["type"]}
            lst_conections.append(d_row )
                
        elif sector["type"] == "single_conductive":
            if sector["way_grounded"] == "not":
                v = 1 #sector["point_grounded"]+1
            elif sector["way_grounded"] in ["left", "right"]:
                v = sector["point_grounded"]
            elif sector["way_grounded"] == "both":
                v = sector["point_grounded"]-1
            elif sector["way_grounded"] == "inside":
                v = sector["point_grounded"]+1
            dl = sector["length"]/v
            #print(dl)

            if sector["countercable"]:
                Y_cc = 1/R_cc(pz,dl,sector["D_countercable"],sector["H_countercable"])

            else:
                Y_cc = 0

            

            for m in range(v):
                if ind == 0 and m==0 and n in nodes_position:
                    i_old = nodes_position[n][0]
                    start = (i_old,j,-1)
                else:
                    start = (i,j,-1)
                i += 1
                end = (i,j,1)
                j += 1
                d_row = {
                    "start": start,
                    "end": end,
                    "type": sector["type"],
                    'supportN': m,
                    'supportK': m+1,
                    "link_okgt": sector["name"],
                    "length":dl,
                    "groundwire": sector["groundwire"], 
                    "H_cable":sector["H_cable"],
                    "X_cable":sector["X_cable"],
                    "countercable": sector["countercable"]
                    }
                lst_conections.append(d_row)
                if sector["countercable"]:
                    lst_cc_conections.append({"is_vl":False, "s":start,"e":end,"length":dl,"sector":sector,})
                    okgt_to_cc[len(lst_conections)-1] = len(lst_cc_conections)-1

                
                if grounded[m] is not None: 
                    yz = (1/sector["point_grounded"] if grounded[m] in ["Yzy","both","both2"] else 0)+\
                        (Y_cc/2 if grounded[m] in ["Ycc","both"] else 0)+(Y_cc if grounded[m] in ["Ycc2","both2"] else 0)     
                    lst_zy.append({"type":"single","Rzy":yz,"p1":start[0]})

                if grounded[v] is not None and m==v-1: 
                    yz = (1/sector["point_grounded"] if grounded[v] in ["Yzy","both","both2"] else 0)+\
                        (Y_cc/2 if grounded[v] in ["Ycc","both"] else 0)+(Y_cc if grounded[v] in ["Ycc2","both2"] else 0)        
                    lst_zy.append({"type":"single","Rzy":yz,"p1":end[0]})

                #

    if k not in nodes_position:
        nodes_position[k] = (i,j)


#print(nodes_position)
""" for row in lst_zy:
    print(row)
 """
def get_vl_sector_info(vl,branch,supports,dtype):
    #"phases", "conductors", "groundwires", "supports", "PSs"
    f = lambda ni,ki,ns,ks: (ni<ki and ns>=ni and ks<=ki) or (ni>ki and ns<=ni and ks>=ki)
    if dtype["name"] == "conductors":
        for item in vl_info[vl]["conductors"]:
            if item["link_branch"]==branch and f(item["supportN"],item["supportK"],supports[0],supports[1]):
                return {"c1":item["conductor"],"c2":item["conductor"],"c3":item["conductor"]}

    elif dtype["name"] == "phases":
        for item in vl_info[vl]["phases"]:
            if item["link_branch"]==branch and f(item["supportN"],item["supportK"],supports[0],supports[1]):
                if item["phase"][0] == "-":
                    mirrored = True
                    p1,p2,p3 = item["phase"][1:]
                else:
                    mirrored = False
                    p1,p2,p3 = item["phase"]
                return {"p1":p1,"p2":p2,"p3":p3,"mirrored":mirrored}

    elif dtype["name"] =="supports":
        for item in vl_info[vl]["supports"]:
            if item["link_branch"]==branch and f(item["supportN"],item["supportK"],supports[0],supports[1]):
                return {"s1":item["support"],"s2":item["support"],"s3":item["support"],"s4":item["support"],"s5":item["support"]}

    elif dtype["name"] =="groundwires":
        for item in vl_info[vl]["groundwires"]:
            if item["link_branch"]==branch and f(item["supportN"],item["supportK"],supports[0],supports[1]):
                if item["type"] == "one":
                    if item["is_okgt"] is None:
                        return {"gw1":item["groundwire1"],"gw2":None,"is_okgt":None}
                    elif item["is_okgt"] is not None and item["is_okgt"]=="groundwire1":
                        return {"gw1":item["groundwire1"],"gw2":None, "is_okgt":"gw1"}
                elif item["type"] == "two":
                    if item["is_okgt"] is None:
                        return {"gw1":item["groundwire1"],"gw2":item["groundwire2"],"is_okgt":None}
                    elif item["is_okgt"] is not None and item["is_okgt"]=="groundwire1":
                        return {"gw1":item["groundwire1"],"gw2":item["groundwire2"],"is_okgt":"gw1"}
                    elif item["is_okgt"] is not None and item["is_okgt"]=="groundwire2":
                        return {"gw1":item["groundwire1"],"gw2":item["groundwire2"],"is_okgt":"gw2"}
        else:
            return {"gw1":None,"gw2":None,"is_okgt":None}

    elif dtype["name"] =="PSs":
        for item in vl_info[vl]["PSs"]:
            if vl_info[vl]["branches"][branch][dtype["side"]] == item["PS_name"]:
                return {"length":item["length"]}

    #"countercables"
    elif dtype["name"] =="countercables":
        for item in vl_info[vl]["countercables"]:
            if item["link_branch"]==branch and f(item["supportN"],item["supportK"],supports[0],supports[1]):
                return item
        else:
            return None

    raise Exception("Supports %s don't exist in %s %s %s" % (str(supports),vl,str(branch),dtype["name"]))


# Construct graph of transmission lines' links
lst_vl_conections = []
i += 1

to_vl = {}

s_i_vl, s_j_vl = i,j

for vl_name, vl in vl_info.items():
    nodes_position = {}
    for key, branch in vl["branches"].items():
        if key[0] not in nodes_position:
            nodes_position[key[0]] = (i,j)

        add_Ps = False
        if branch["PS"] in ["both", "left"]:
            add_Ps = True
            nodes_position[key[0]] = (i,j)
        trig_first = False
        order = sector_entry(vl["sectors"],key)
        #print(key, order)
        for sector in [vl["sectors"][i] for i in order]:
            #print(key, sector["link_branch"])
            if key == sector["link_branch"]:
                #print("aaaaaaaaaa")
                N, K = sector["supportN"], sector["supportK"]
                drct = int((K-N)/abs(K-N))
                gen = zip(range(abs(K-N)),range(N,K,drct),range(N+drct,K+drct,drct))
                

                if sector["type"]=="with_okgt":
                    dl = abs(sector['lengthK']-sector['lengthN'])/abs(K-N)
                elif sector["type"]=="without_okgt":
                    dl = abs(sector['length'])/abs(K-N)

                for t, nc, kc in gen:
                    data1,data2,data3,data4,data5 = {},{},{},{},{}

                    keyword = ["name_vl","supportN","supportK","link_branch"]
                    dt = [vl_name, nc, kc, key]
                    for kw,rz in zip(keyword,dt):
                        data1[kw],data2[kw],data3[kw],data4[kw],data5[kw] = rz,rz,rz,rz,rz

                    #"phases", "conductors", "groundwires", "supports"
                    phases = get_vl_sector_info(vl_name,key,(nc, kc,),{"name":"phases"})
                    conductors = get_vl_sector_info(vl_name,key,(nc, kc,),{"name":"conductors"})
                    supports = get_vl_sector_info(vl_name,key,(nc, kc,),{"name":"supports"})
                    groundwires = get_vl_sector_info(vl_name,key,(nc, kc,),{"name":"groundwires"})
                    countercables = get_vl_sector_info(vl_name,key,(nc, kc,),{"name":"countercables"})


                    keyword = ["type","is_Ps_sector","length","Tsupport","phase","mirrored","conductor"] #"start","end",
                    d1 = [sector["type"],False,dl,supports["s1"],phases["p1"],phases["mirrored"],conductors["c1"]] #s1,e1,
                    d2 = [sector["type"],False,dl,supports["s2"],phases["p2"],phases["mirrored"],conductors["c2"]] #s2,e2,
                    d3 = [sector["type"],False,dl,supports["s3"],phases["p3"],phases["mirrored"],conductors["c3"]] #s3,e3,
                    d4 = [sector["type"],False,dl,supports["s4"],"T1",phases["mirrored"],[groundwires["gw1"],groundwires["is_okgt"]=="gw1"]] #s4,e4,
                    d5 = [sector["type"],False,dl,supports["s5"],"T2",phases["mirrored"],[groundwires["gw2"],groundwires["is_okgt"]=="gw2"]] #s5,e5,

                    for kw,rz1,rz2,rz3,rz4,rz5 in zip(keyword,d1,d2,d3,d4,d5):
                        data1[kw],data2[kw],data3[kw],data4[kw],data5[kw] = rz1,rz2,rz3,rz4,rz5

                    if not trig_first and t==0 and key[0] in nodes_position:
                        if not add_Ps:
                            trig_first = True
                            i_old = nodes_position[key[0]][0]
                            s1,s2,s3,s4,s5 = (i_old,j,-1),(i_old+1,j+1,-1),(i_old+2,j+2,-1),(i_old+3,j+3,-1),(i_old+4,j+4,-1)
                            i+=5
                            e1,e2,e3,e4,e5= (i,j,1),(i+1,j+1,1),(i+2,j+2,1),(i+3,j+3,1),(i+4,j+4,1)
                        else:
                            trig_first = True
                            add_Ps = False
                            i_old = nodes_position[key[0]][0]
                            s1,s2,s3,s4,s5 = (i_old,j,-1),(i_old+1,j+1,-1),(i_old+2,j+2,-1),(i_old+3,j+3,-1),(i_old+4,j+4,-1)
                            i+=5
                            e1,e2,e3,e4,e5= (i,j,1),(i+1,j+1,1),(i+2,j+2,1),(i+3,j+3,1),(i+4,j+4,1)
                            j += 5
                            ps = get_vl_sector_info(vl_name,key,(nc, kc,),{"name":"PSs","side":"PS_name_1"})
                            lst_vl_conections.append({"start":s1,"end":e1,**data1,"length":ps["length"],"is_Ps_sector":True})
                            lst_vl_conections.append({"start":s2,"end":e2,**data2,"length":ps["length"],"is_Ps_sector":True})
                            lst_vl_conections.append({"start":s3,"end":e3,**data3,"length":ps["length"],"is_Ps_sector":True})
                            lst_vl_conections.append({"start":s4,"end":e4,**data4,"length":ps["length"],"is_Ps_sector":True})
                            lst_vl_conections.append({"start":s5,"end":e5,**data5,"length":ps["length"],"is_Ps_sector":True})
                            
                            s1,s2,s3,s4,s5 = (i,j,-1),(i+1,j+1,-1),(i+2,j+2,-1),(i+3,j+3,-1),(i+4,j+4,-1)
                            i += 5
                            e1,e2,e3,e4,e5= (i,j,1),(i+1,j+1,1),(i+2,j+2,1),(i+3,j+3,1),(i+4,j+4,1)
                            

                    else:
                        s1,s2,s3,s4,s5 = (i,j,-1),(i+1,j+1,-1),(i+2,j+2,-1),(i+3,j+3,-1),(i+4,j+4,-1)
                        i += 5
                        e1,e2,e3,e4,e5= (i,j,1),(i+1,j+1,1),(i+2,j+2,1),(i+3,j+3,1),(i+4,j+4,1)

                    
                    #e1,e2,e3,e4,e5= (i,j,1),(i+1,j+1,1),(i+2,j+2,1),(i+3,j+3,1),(i+4,j+4,1)
                    j += 5

                    if sector["type"]=="with_okgt":
                        to_vl[(vl_name,key,nc,kc)] = len(lst_vl_conections)

                    lst_vl_conections.append({"start":s1,"end":e1,**data1})
                    lst_vl_conections.append({"start":s2,"end":e2,**data2})
                    lst_vl_conections.append({"start":s3,"end":e3,**data3})
                    lst_vl_conections.append({"start":s4,"end":e4,**data4})
                    lst_vl_conections.append({"start":s5,"end":e5,**data5})

                    if countercables is not None:
                        lst_cc_conections.append({"is_vl":True, "vl_name":vl_name,"branch":key,"N":nc,"K":kc, "s":s4,"e":e4,"length":dl,"sector":countercables})


        if branch["PS"] in ["both", "right"]:
            s1,s2,s3,s4,s5 = (i,j,-1),(i+1,j+1,-1),(i+2,j+2,-1),(i+3,j+3,-1),(i+4,j+4,-1)
            i += 5
            e1,e2,e3,e4,e5= (i,j,1),(i+1,j+1,1),(i+2,j+2,1),(i+3,j+3,1),(i+4,j+4,1)
            ps = get_vl_sector_info(vl_name,key,(nc, kc,),{"name":"PSs","side":"PS_name_2"})
            lst_vl_conections.append({"start":s1,"end":e1,**data1,"length":ps["length"],"is_Ps_sector":True})
            lst_vl_conections.append({"start":s2,"end":e2,**data2,"length":ps["length"],"is_Ps_sector":True})
            lst_vl_conections.append({"start":s3,"end":e3,**data3,"length":ps["length"],"is_Ps_sector":True})
            lst_vl_conections.append({"start":s4,"end":e4,**data4,"length":ps["length"],"is_Ps_sector":True})
            lst_vl_conections.append({"start":s5,"end":e5,**data5,"length":ps["length"],"is_Ps_sector":True})
            
            i+=5
            j+=5
            

        if key[1] not in nodes_position:
            nodes_position[key[1]] = (i,j)


s_i_сс, s_j_сс = i,j
cc_lst_conections = []
to_cc = {}
for item in lst_cc_conections:
    if item["is_vl"]:
        to_cc[(item["vl_name"],item["branch"],item["N"],item["K"])] = len(cc_lst_conections)
    d = {"is_vl":item["is_vl"],
        "start":(item["s"][0],j,-1),
        "end":(item["e"][0],j,1),
        "link_s":item["s"],
        "link_e":item["e"],
        "length":item["length"],
        "H_countercable":item["sector"]["H_countercable"],
        "X_countercable": item["sector"]["X_countercable"],
        "D_countercable": item["sector"]["D_countercable"],
        "ground_resistance": item["sector"]["ground_resistance"],
        }
    cc_lst_conections.append(d)
    
    j+=1

s_i_end, s_j_end = i,j
    

""" for idd,row in enumerate(lst_vl_conections):
    print(row)
    if idd>100:break """



okgt_to = {}
vl_to = {}
cc_to = {}

for i,item in enumerate(lst_conections):
    if item["type"]=="VL":
        key = (item['name_vl'],item['link_branch'],item['supportN'],item['supportK'])
        pos = (*item['start'][:2],*item['end'][:2])
        cc = (*cc_lst_conections[to_cc[key]]["start"][:2],*cc_lst_conections[to_cc[key]]["end"][:2]) if key in to_cc else None
        vl = (*lst_vl_conections[to_vl[key]]["start"][:2],*lst_vl_conections[to_vl[key]]["end"][:2])
        okgt_to[pos] = {"vl":(vl,to_vl[key]),"cc":(cc,to_cc.get(key))}
        vl_to[vl] = {"okgt":(pos,i),"cc":(cc,to_cc.get(key))}
        if cc is not None:
            cc_to[cc] = {"okgt":(pos,i),"vl":(vl,to_vl[key])}
        
    elif item["type"]=="single_conductive" and i in okgt_to_cc:
        pos = (*item['start'][:2],*item['end'][:2])
        cc = (*cc_lst_conections[okgt_to_cc[i]]['start'][:2],*cc_lst_conections[okgt_to_cc[i]]['end'][:2])
        okgt_to[pos] = {"vl":(None,None),"cc":(cc,okgt_to_cc[i])}
        cc_to[cc] = {"okgt":(pos,i),"vl":(None,None)}
        
    """ elif item["type"]=='single_dielectric':
        pass """


A = sparse.lil_matrix((s_i_end, s_j_end),dtype=np.float64)

for item in lst_conections: 
    A[item["start"][:2]] = item["start"][2]
    A[item["end"][:2]] = item["end"][2]

for item in lst_vl_conections:
    A[item["start"][:2]] = item["start"][2]
    A[item["end"][:2]] = item["end"][2]

for item in cc_lst_conections:
    A[item["start"][:2]] = item["start"][2]
    A[item["end"][:2]] = item["end"][2]


pz = 30
carson_cashe = {}
mu0=np.pi*4*10**(-7)
jp=1j*2*np.pi*50
sh_int=np.arange(0, 1, 0.00001)
mp=(mu0*jp*(1/pz))**0.5
ro_fe =  0.1 #Ом*мм2/м
R_isol = 10**8

#R_cc = lambda d:(4*ro_fe)/(np.pi*d**2)
Cars_ii = lambda x,hi,ri:(np.exp(-2*x*hi)*np.cos(x*ri))/(x+(x**2+mp**2)**0.5)
Cars_ij = lambda x,hi,hj,xi,xj:(np.exp(-x*(hi+hj))*np.cos(x*abs(xi-xj)))/(x+(x**2+mp**2)**0.5)

def Carson(data):
    dd = tuple([tuple(data[i]) if i<5 else data[i] for i in range(7)])
    if dd in carson_cashe:
        return carson_cashe[dd]
    else:
        X,Y,R,r,trig,pz,l = dd 
        mp=(mu0*jp*(1/pz))**0.5
        

        size = len(X)
        M = sparse.lil_matrix((size,size), dtype=np.complex128)

        for a in range(size):
            for b in range(a,size):
                if a == b and trig[a]:
                    M[a,b]=R[a]*l + jp*mu0/2/np.pi*(np.log(2*abs(Y[a])/r[a])+2*simps(Cars_ii(sh_int,Y[a],r[a]),sh_int))*l*10**3
                elif a == b and not trig[a]:
                    M[a,b]=R[a]
                elif a!=b and (not trig[a] or not trig[b]):
                    M[a,b]=M[b,a]=0
                elif a!=b and not trig[a] and not trig[b]:
                    M[a,b]=M[b,a]=jp*mu0/4/np.pi*(np.log(((Y[a]+Y[b])**2+(X[a]-X[b])**2)/((Y[a]-Y[b])**2+(X[a]-X[b])**2))\
                        +4*simps(Cars_ij(sh_int,Y[a],Y[b],X[a],X[b]),sh_int))*l*10**3


        carson_cashe[dd] = M
        return M
        

def init_carson_data(key,data,shift):
    if key=="phase":
        support = k_supports[data["Tsupport"]]
        mirrored = -1 if data["mirrored"] else 1
        X = mirrored*support["X"][data["phase"]]
        Y = support["Y"][data["phase"]]-support["isulator"]+shift
        R = k_conductors[data["conductor"]]["R0"]
        r = k_conductors[data["conductor"]]["r"]
        return X,Y,R,r
        
    elif key=="groundwire":
        if data["conductor"][1] or (not data["conductor"][1] and data["conductor"][0] is None):
            return None,None,R_isol,None
        elif not data["conductor"][1] and data["conductor"][0] is not None:
            support = k_supports[data["Tsupport"]]
            mirrored = -1 if data["mirrored"] else 1
            X = mirrored*support["X"][data["phase"]]
            Y = support["Y"][data["phase"]]
            if Y==0 and X==0:
                raise Exception("%s doesn't support groundwire %s" % (data["Tsupport"],data["phase"]))
            else:
                Y+=shift
            R = k_conductors[data["conductor"][0]]["R0"]
            r = k_conductors[data["conductor"][0]]["r"]
            return X,Y,R,r

    elif key=="okgt_vl":
        support = k_supports[data["Tsupport"]]
        mirrored = -1 if data["mirrored"] else 1
        X = mirrored*support["X"][data["phase"]]
        Y = support["Y"][data["phase"]]
        if Y==0 and X==0:
            raise Exception("%s doesn't support groundwire %s" % (data["Tsupport"],data["phase"]))
        else:
            Y+=shift
        R = k_conductors[data["conductor"][0]]["R0"]
        r = k_conductors[data["conductor"][0]]["r"]
        return X,Y,R,r
    elif key=="cc":
        X = data["X_countercable"]
        Y = data["H_countercable"]+shift
        r = data["D_countercable"]/2*10**-3
        R = (4*ro_fe)/(np.pi*data["D_countercable"]**2)*10**3
        return X,Y,R,r

    elif key=="okgt_single":
        X = data['X_cable']
        Y = data['H_cable']+shift
        R = k_conductors[data['groundwire']]["R0"]
        r = k_conductors[data['groundwire']]["r"]
        return X,Y,R,r

    elif key=="phase_ps":
        return None,None,R_isol,None
    elif key=="groundwire_ps":
        if data["conductor"][0] is None:
            return None,None,R_isol,None
        else:
            support = k_supports[data["Tsupport"]]
            mirrored = -1 if data["mirrored"] else 1
            X = mirrored*support["X"][data["phase"]]
            Y = support["Y"][data["phase"]]
            if Y==0 and X==0:
                raise Exception("%s doesn't support groundwire %s" % (data["Tsupport"],data["phase"]))
            else:
                Y+=shift
            R = k_conductors[data["conductor"][0]]["R0"]
            r = k_conductors[data["conductor"][0]]["r"]
            return X,Y,R,r


Z = sparse.lil_matrix((s_j_end, s_j_end),dtype=np.complex128)
for i in range(0,len(lst_vl_conections),5):
    item = lst_vl_conections[i]
    if item["type"]=="with_okgt" and not item['is_Ps_sector'] and\
        vl_to[(*item["start"][:2],*item["end"][:2])]["cc"][0] is not None:
        vl_d = lst_vl_conections[i:i+5]
        cc = cc_lst_conections[vl_to[(*item["start"][:2],*item["end"][:2])]["cc"][1]]
        okgt = lst_conections[vl_to[(*item["start"][:2],*item["end"][:2])]["okgt"][1]]

        shift = cc["H_countercable"]+0.1

        t1 = vl_d[3]["conductor"]
        t2 = vl_d[4]["conductor"]
        pos = 3 if t1[1] else 4
        
        dl = vl_d[0]["length"]

        trig = [True,True,True,(not t1[1] and t1[0] is not None),(not t2[1] and t2[0] is not None),True,True]

        k = ["phase","phase","phase","groundwire","groundwire","okgt_vl","cc"]
        d = [vl_d[0],vl_d[1],vl_d[2],vl_d[3],vl_d[4],vl_d[pos],cc]

        X,Y,R,r = [],[],[],[]

        for key, data in zip(k,d):
            Xi,Yi,Ri,ri = init_carson_data(key,data,shift)
            X.append(Xi)
            Y.append(Yi)
            R.append(Ri)
            r.append(ri)
        
        M = Carson([X,Y,R,r,trig,pz,dl])
        ivln,ivlk = vl_d[0]["start"][1],vl_d[4]["start"][1]
        iokgt = okgt["start"][1]
        icc = cc["start"][1]

        Z[ivln:ivlk+1,ivln:ivlk+1] = M[:5,:5]
        Z[iokgt,iokgt] = M[5,5]
        Z[icc,icc] = M[6,6]

        Z[ivln:ivlk+1,iokgt] = M[:5,5]
        Z[ivln:ivlk+1,icc] = M[:5,6]

        Z[iokgt,ivln:ivlk+1] = M[5,:5]
        Z[icc,ivln:ivlk+1] = M[6,:5]

        Z[iokgt,icc] = M[5,6]
        Z[icc,iokgt] = M[6,5]


    elif item["type"]=="with_okgt" and not item['is_Ps_sector'] and\
        vl_to[(*item["start"][:2],*item["end"][:2])]["cc"][0] is None:
        vl_d = lst_vl_conections[i:i+5]
        okgt = lst_conections[vl_to[(*item["start"][:2],*item["end"][:2])]["okgt"][1]]

        shift = 0

        t1 = vl_d[3]["conductor"]
        t2 = vl_d[4]["conductor"]
        pos = 3 if t1[1] else 4
        
        dl = vl_d[0]["length"]

        trig = [True,True,True,(not t1[1] and t1[0] is not None),(not t2[1] and t2[0] is not None),True]

        k = ["phase","phase","phase","groundwire","groundwire","okgt_vl"]
        d = [vl_d[0],vl_d[1],vl_d[2],vl_d[3],vl_d[4],vl_d[pos]]

        X,Y,R,r = [],[],[],[]

        for key, data in zip(k,d):
            Xi,Yi,Ri,ri = init_carson_data(key,data,shift)
            X.append(Xi)
            Y.append(Yi)
            R.append(Ri)
            r.append(ri)
        
        M = Carson([X,Y,R,r,trig,pz,dl])
        ivln,ivlk = vl_d[0]["start"][1],vl_d[4]["start"][1]
        iokgt = okgt["start"][1]

        Z[ivln:ivlk+1,ivln:ivlk+1] = M[:5,:5]
        Z[iokgt,iokgt] = M[5,5]

        Z[ivln:ivlk+1,iokgt] = M[:5,5]
        Z[iokgt,ivln:ivlk+1] = M[5,:5]
        

    elif item["type"]=="without_okgt" and not item['is_Ps_sector']:
        vl_d = lst_vl_conections[i:i+5]

        shift = 0

        t1 = vl_d[3]["conductor"]
        t2 = vl_d[4]["conductor"]
        pos = 3 if t1[1] else 4
        
        dl = vl_d[0]["length"]

        trig = [True,True,True,(not t1[1] and t1[0] is not None),(not t2[1] and t2[0] is not None)]

        k = ["phase","phase","phase","groundwire","groundwire"]
        d = [vl_d[0],vl_d[1],vl_d[2],vl_d[3],vl_d[4]]

        X,Y,R,r = [],[],[],[]

        for key, data in zip(k,d):
            Xi,Yi,Ri,ri = init_carson_data(key,data,shift)
            X.append(Xi)
            Y.append(Yi)
            R.append(Ri)
            r.append(ri)
        
        M = Carson([X,Y,R,r,trig,pz,dl])
        ivln,ivlk = vl_d[0]["start"][1],vl_d[4]["start"][1]
        
        Z[ivln:ivlk+1,ivln:ivlk+1] = M[:5,:5]

    elif (item["type"]=="with_okgt" or item["type"]=="without_okgt") and item['is_Ps_sector']:
        vl_d = lst_vl_conections[i:i+5]

        shift = 0

        t1 = vl_d[3]["conductor"]
        t2 = vl_d[4]["conductor"]
        pos = 3 if t1[1] else 4
        
        dl = vl_d[0]["length"]

        trig = [False,False,False,t1[0] is not None,t2[0] is not None]

        k = ["phase_ps","phase_ps","phase_ps","groundwire_ps","groundwire_ps"]
        d = [vl_d[0],vl_d[1],vl_d[2],vl_d[3],vl_d[4]]

        X,Y,R,r = [],[],[],[]

        for key, data in zip(k,d):
            Xi,Yi,Ri,ri = init_carson_data(key,data,shift)
            X.append(Xi)
            Y.append(Yi)
            R.append(Ri)
            r.append(ri)
        
        M = Carson([X,Y,R,r,trig,pz,dl])
        ivln,ivlk = vl_d[0]["start"][1],vl_d[4]["start"][1]
        
        Z[ivln:ivlk+1,ivln:ivlk+1] = M[:5,:5]


for i, item in enumerate(lst_conections):
    if item["type"]=="VL":
        continue
    elif item["type"]=='single_conductive':
        if item['countercable']:
            pos = (*item['start'][:2],*item['end'][:2])
            cc = cc_lst_conections[okgt_to[pos]["cc"][1]]
            shift = cc["H_countercable"]+0.1
            dl = item["length"]

            trig = [True,True]
            k = ["okgt_single","cc"]
            d = [item,cc]
            X,Y,R,r = [],[],[],[]

            for key, data in zip(k,d):
                Xi,Yi,Ri,ri = init_carson_data(key,data,shift)
                X.append(Xi)
                Y.append(Yi)
                R.append(Ri)
                r.append(ri)

            M = Carson([X,Y,R,r,trig,pz,dl])
            iokgt,icc = item["start"][1],cc["start"][1]
            
            Z[iokgt,iokgt] = M[0,0]
            Z[icc,icc] = M[1,1]

            Z[iokgt,icc] = M[0,1]
            Z[iokgt,icc] = M[0,1]
        else:
            shift = 0
            dl = item["length"]

            trig = [True]
            k = ["okgt_single"]
            d = [item]
            X,Y,R,r = [],[],[],[]

            for key, data in zip(k,d):
                Xi,Yi,Ri,ri = init_carson_data(key,data,shift)
                X.append(Xi)
                Y.append(Yi)
                R.append(Ri)
                r.append(ri)
            
            M = Carson([X,Y,R,r,trig,pz,dl])
            iokgt = item["start"][1]
            
            Z[iokgt,iokgt] = M[0,0]
        
    elif item["type"]=='single_dielectric':
        X,Y,R,r,trig = [None],[None],[R_isol],[None],[False]
        M = Carson([X,Y,R,r,trig,pz,0])
        iokgt = item["start"][1]
        Z[iokgt,iokgt] = M[0,0]
        
for i in range(s_j_end):
    if Z[i,i] == 0:
        print(i)     

print("end",s_j_end)