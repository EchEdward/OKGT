# pylint: disable=E0611
# pylint: disable=E1101
#pyi-makespec --onefile --icon=icon.ico --noconsole VEZRead.py
from PyQt5.QtWidgets import QApplication,QMainWindow,QWidget,QVBoxLayout,QHBoxLayout,QLabel,\
    QScrollArea,QSizePolicy,QSplitter, QSizePolicy, qApp, QAction,\
    QMessageBox,QFileDialog, QErrorMessage, QDoubleSpinBox, QSpacerItem, QLineEdit, QItemDelegate, QProgressBar,\
    QTabWidget, QComboBox, QGridLayout, QCheckBox, QSpinBox, QDoubleSpinBox, QSpacerItem, QProgressDialog,\
    QButtonGroup, QRadioButton, QTextEdit

from PyQt5.QtGui import QPixmap, QPalette, QBrush, QImage, QIcon, QTransform, QStandardItemModel,QStandardItem,\
     QDoubleValidator, QValidator, QCloseEvent, QColor
from PyQt5.QtCore import QPersistentModelIndex, Qt,  QSize, QModelIndex, QThread, pyqtSignal, QTimer

import sys
import os

from table_classes import traceback_erors, OkgtSectorTable, OkgtSingleTable, PSTable, NodeTable, LineEditManager,\
    VlParamsTable, VlSectorTable, VlPsParamsTable, VlCommonChainsTable, RPASettingsTable, ShortCircuitLineTable,\
    UserComboBox, CustomDialog


from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
from matplotlib.ticker import FuncFormatter
import matplotlib.pyplot as plt

from openpyxl import load_workbook

import json

from calc_okgt import k_supports, k_conductors, main_calc, I_sc_corector

from initial_data2 import okgt_info, vl_info, ps_info, rpa_info

from report_creator import memorandum, explanatory

import traceback





class OkgtCalcTread(QThread):
    mysignal = pyqtSignal(str)
    def __init__(self, parent=None):
        QThread.__init__(self, parent)
        self.okgt_info = None
        self.vl_info = None
        self.ps_info = None
        self.rpa_info = None


    def setData(self, okgt_info, vl_info, ps_info, rpa_info):
        self.okgt_info = okgt_info
        self.vl_info = vl_info
        self.ps_info = ps_info
        self.rpa_info = rpa_info
        
    def run(self):
        self.mysignal.emit('start:')
        try:
            self.result = main_calc(self.okgt_info, self.vl_info, self.ps_info, self.rpa_info, callback=self.runningMesages)
        except Exception as ex:
            self.mysignal.emit(f'error:{ex}')
        else:
            self.mysignal.emit('end:')


    def getData(self):
        return self.result

    def runningMesages(self,tp,adds):
        if tp == "Calc equation system":
            self.mysignal.emit('run_messages:Составляем систему уравнений')
        elif tp == "Calc points":
            addsn = (adds[0],adds[1],adds[2],f"{adds[3][0]}-{adds[3][1]}",adds[4])
            self.mysignal.emit('run_points:{"current":"%s","length":"%s","vl_name":"%s","branch":"%s","support":"%s"}' % addsn)

            
    # остановка потока
    def stop( self ):
        
        self.terminate()
        self.wait()
        return 'stop'
    

class MyWindow(QMainWindow):
    def __init__(self,parent=None):
        super(MyWindow,self).__init__(parent)

        self.setSettings()

        self.currentFilePath = ''
        self.currentFileName = 'Новый файл'
        self.setWindowTitle(f"OKGT - {self.currentFileName}")
        self.setWindowIcon(QIcon("icon.ico"))

        desktop = QApplication.desktop()
        wd = desktop.width()
        hg = desktop.height()
        ww = 1000
        wh = 500
        if ww>wd: ww = int(0.7*wd)
        if wh>hg: wh = int(0.7*hg)
        x = (wd-ww)//2
        y = (hg-wh)//2
        self.setGeometry(x, y, ww, wh)

        topVBoxLayout = QVBoxLayout(self)
        self.mainTabWidget = QTabWidget()
        topVBoxLayout.addWidget(self.mainTabWidget) 
        central_widget = QWidget()
        central_widget.setLayout(topVBoxLayout)
        self.setCentralWidget(central_widget) 

        
        self.Okgt_Widget = QWidget()
        self.Vl_Tabs = QTabWidget()
        self.Ps_Widget = QWidget()
        self.Rpa_Tabs = QTabWidget()
        self.Rez_Tabs = QTabWidget()
        self.Rez_Tabs = QTabWidget()
        self.Report_Widget = QWidget()

        self.mainTabWidget.addTab(self.Okgt_Widget,"ОКГТ")
        self.mainTabWidget.addTab(self.Ps_Widget,"ПС")
        self.mainTabWidget.addTab(self.Vl_Tabs,"ВЛ")
        self.mainTabWidget.addTab(self.Rpa_Tabs,"РЗА")
        self.mainTabWidget.addTab(self.Rez_Tabs,"Результаты")
        self.mainTabWidget.addTab(self.Report_Widget,"Настроики документов")


        self.cntr_pr = {"trig":False,"timer":QTimer()}
        self.cntr_pr["timer"].timeout.connect(self.on_timeout)
        self.cntr_pr["timer"].setSingleShot(True)


        self.okgt_tab_maker()
        self.ps_tab_maker()
        self.MenuBarMaker()
        self.ReportTab()

        self.le_manager = LineEditManager(self.Vl_Tabs)
        self.vl_liks = {}
        self.rpa_liks = {}
        self.vl_settings_dict = {
            "conductors":"Провода","phases":"Фазировка","supports":"Опоры","groundwires":"Грозотросы",\
            "PSs":"ПС","grounded":"Заземление проводов","countercables":"Противовес","commonchains":"Смежные цепи",
        }
        self.vl_settings_dict_rev = {v:k for k,v in self.vl_settings_dict.items()}

        self.okgt_calc_tread = OkgtCalcTread()
        self.okgt_calc_tread.mysignal.connect(self.OkgtCalculationSignals, Qt.QueuedConnection)
        self.okgt_proces_dialog = None

        self.resFig = {}
        self.jsSpleater = "\n\u00C6\n"
        self.font_s = 14

        self.calc_results = None
        self.sectorsFig = {}

    


    def setSettings(self):
        try:
            self.path_home = os.path.expanduser("~\\Desktop\\")
        except Exception:
            self.path_home = ""

        try:
            current_os = sys.platform
            if current_os == "win32":
                self.path_midle_files = os.path.expanduser('~\\AppData\\Roaming\\')
            elif current_os.startswith('linux'):
                self.path_midle_files = os.path.expanduser('~\\')
            elif current_os == 'darwin':
                self.path_midle_files = os.path.expanduser('~\\')
        except Exception:
            self.path_midle_files = ''

        

        for some_dir in ['okgt','okgt\\figures']:
            curren_dir = os.path.join(self.path_midle_files,some_dir)
            if (os.path.exists(curren_dir) and not os.path.isdir(curren_dir)) or not os.path.exists(curren_dir):
                try:
                    os.mkdir(curren_dir)
                except OSError:
                    print ("Error generate dir "+curren_dir)

        self.path_midle_files = os.path.join(self.path_midle_files,'okgt')

        last_path_keys = ['id_save_as','id_open','memorandum_path','explanatory_path','excel_open']
        report_settings = {
            "show_arc_pause":True,
            "show_Irpa":True,
            "recipients":'',
            "department_boss_type":False,
            "department_boss_name":'',
            "group_boss_name":'',
        }
        try: 
            with open(os.path.join(self.path_midle_files,'main_settings.json'), "r" ) as f:
                self.main_settings = {i:j for i,j in json.load(f).items()}
            
        except Exception:
            self.main_settings = {}
        finally:
            for path_key in last_path_keys:
                if path_key not in self.main_settings:
                    self.main_settings[path_key] = self.path_home

            for setting, val in report_settings.items():
                if setting not in self.main_settings:
                    self.main_settings[setting] = val    

    

    def on_timeout(self):
        self.cntr_pr["trig"] = False

    def keyPressEvent(self, e):
        if e.key() == Qt.Key_Control:
            self.cntr_pr["trig"] = True
            self.cntr_pr["timer"].start(500)
            
    def keyReleaseEvent(self,e):
        if e.key() == Qt.Key_Control:
            self.cntr_pr["trig"] = False


    
    def ReportTab(self):
        Vlayout = QVBoxLayout()

        show_arc_pause = QCheckBox('Отображать длительность паузы между откл. КЗ и АПВ')
        Vlayout.addWidget(show_arc_pause)
        show_Irpa = QCheckBox('Отображать токи уставок РЗА')
        Vlayout.addWidget(show_Irpa)

        Vlayout.addWidget(QLabel('Получатели служебной записки:'))
        recipients = QLineEdit()
        #recipients.setMaximumHeight(70)
        Vlayout.addWidget(recipients)

        Vlayout.addWidget(QLabel('Название проекта:'))
        project_name = QTextEdit()
        project_name.setMaximumHeight(70)
        Vlayout.addWidget(project_name)

        rb1 = QRadioButton('Начальник')
        rb2 = QRadioButton('Зам. начальника')

        button_group = QButtonGroup()
        button_group.addButton(rb1,0)
        button_group.addButton(rb2,1)

        type_boss_layout = QHBoxLayout()
        type_boss_layout.addWidget(rb1)
        type_boss_layout.addWidget(rb2)
        Vlayout.addLayout(type_boss_layout)

        department_boss_name = QLineEdit()
        Vlayout.addWidget(department_boss_name)

        Vlayout.addWidget(QLabel('Заведующий группы:'))
        group_boss_name = QLineEdit()
        Vlayout.addWidget(group_boss_name)

        Vlayout.addStretch(2)

        Hlayout = QHBoxLayout()
        Hlayout.addLayout(Vlayout)
        Hlayout.addStretch(2)
        Hlayout.setStretchFactor(Vlayout,1)
        

        self.Report_Widget.setLayout(Hlayout)

        self.ReportSetingsForm = {
            "show_arc_pause":show_arc_pause,
            "show_Irpa":show_Irpa,
            "recipients":recipients,
            "project_name":project_name,
            "department_boss_type":button_group,
            "department_boss_name":department_boss_name,
            "group_boss_name":group_boss_name,
        }

        self.setReportSettings(self.main_settings)
        

    def setReportSettings(self, source):
        self.ReportSetingsForm["show_arc_pause"].setCheckState(Qt.Checked if source["show_arc_pause"] else Qt.Unchecked)
        self.ReportSetingsForm["show_Irpa"].setCheckState(Qt.Checked if source["show_Irpa"] else Qt.Unchecked)
        self.ReportSetingsForm["recipients"].setText(source["recipients"])
        self.ReportSetingsForm["department_boss_name"].setText(source["department_boss_name"])
        self.ReportSetingsForm["group_boss_name"].setText(source["group_boss_name"])
        self.ReportSetingsForm["department_boss_type"].button(int(source["department_boss_type"])).setChecked(True)

        if "project_name" in source:
            self.ReportSetingsForm["project_name"].setText(source["project_name"])

    @traceback_erors
    def getReportSettings(self):
        return {
            "show_arc_pause": True if self.ReportSetingsForm["show_arc_pause"].checkState()==Qt.Checked else False,
            "show_Irpa": True if self.ReportSetingsForm["show_Irpa"].checkState()==Qt.Checked else False,
            "recipients": self.ReportSetingsForm["recipients"].text(),
            "project_name": self.ReportSetingsForm["project_name"].toPlainText(),
            "department_boss_name": self.ReportSetingsForm["department_boss_name"].text(),
            "group_boss_name": self.ReportSetingsForm["group_boss_name"].text(),
            "department_boss_type": bool(self.ReportSetingsForm["department_boss_type"].checkedId())
        }
        
    
    def SaveReportSettings(self):
        try:
            for key, val in self.getReportSettings().items():
                if key != "project_name":
                    self.main_settings[key] = val
        except Exception:
            pass
        else:
            QMessageBox.information(self, 'Сохранение параметров отчёта',
                                        'Операция прошла успешно.',
                                        buttons=QMessageBox.Ok,
                                        defaultButton=QMessageBox.Ok)

    
    def CreateMemorandumDoc(self):
        try:
            fname = QFileDialog.getSaveFileName(self, 'Сохранить служебную записку',\
                 os.path.join(self.main_settings['memorandum_path'],self.currentFileName+' Служебная записка'),'*.docx')
            if fname[0] == "" and fname[1] == "": return
            fname = fname[0]
            self.main_settings['memorandum_path'] = os.path.dirname(fname)
            #self.calc_results = None
            #self.sectorsFig = {}

            report_setings = self.getReportSettings()
            okgt_info_new, _, vl_info_new,  rpa_info_new = self.ReadTables()

            

            memorandum(fname, self.path_midle_files, okgt_info_new, vl_info_new, rpa_info_new, self.calc_results, report_setings)
            
        except Exception as ex:
            ems = QErrorMessage(self)
            ems.setWindowTitle('Возникла ошибка')
            ems.showMessage('Не получилось сохранить файл. '+
                            'Проверьте введённые данные а также сохраняете ли вы уже в открытый файл.'+str(ex))
        else:
            QMessageBox.information(self, 'Сохранение служебной записки','Операция прошла успешно.',
                                          buttons=QMessageBox.Ok,
                                          defaultButton=QMessageBox.Ok)


    def CreateExplanatoryDoc(self):
        try:
            fname = QFileDialog.getSaveFileName(self, 'Сохранить пояснительную записку',\
                 os.path.join(self.main_settings['explanatory_path'],self.currentFileName+' Пояснительная записка'),'*.docx')
            if fname[0] == "" and fname[1] == "": return
            fname = fname[0]
            self.main_settings['explanatory_path'] = os.path.dirname(fname)
            #self.calc_results = None
            #self.sectorsFig = {}

            report_setings = self.getReportSettings()
            okgt_info_new, _, vl_info_new,  rpa_info_new = self.ReadTables()

            explanatory(fname, self.path_midle_files, okgt_info_new, vl_info_new, rpa_info_new, report_setings, self.calc_results, self.sectorsFig,self.rpa_liks)
            
        except Exception as ex:
            ems = QErrorMessage(self)
            ems.setWindowTitle('Возникла ошибка')
            ems.showMessage('Не получилось сохранить файл. '+
                            'Проверьте введённые данные а также сохраняете ли вы уже в открытый файл.'+str(ex))
        else:
            QMessageBox.information(self, 'Сохранение пояснительной записки','Операция прошла успешно.',
                                          buttons=QMessageBox.Ok,
                                          defaultButton=QMessageBox.Ok)
        
    

    def MenuBarMaker(self):
        self.statusBar()
        menubar = self.menuBar()

        fileMenu = menubar.addMenu('&Файл')

        open_file = QAction( '&Открыть', self) #QIcon('exit.png'),
        open_file.setShortcut('Ctrl+O')
        open_file.setStatusTip('Открыть файл ИД')
        open_file.triggered.connect(self.InitialDataOpen)
        fileMenu.addAction(open_file)

        id_save = QAction( '&Сохранить', self) #QIcon('exit.png'),
        id_save.setShortcut("Ctrl+Shift+S")
        id_save.setStatusTip('Сохранить исходные данные в текущий файл')
        id_save.triggered.connect(self.InitialDataSave)
        fileMenu.addAction(id_save)

        id_save_as = QAction( '&Сохранить как', self) #QIcon('exit.png'),
        id_save_as.setShortcut('Ctrl+P')
        id_save_as.setStatusTip('Сохранить исходные данные в новый файл')
        id_save_as.triggered.connect(self.InitialDataSaveAs)
        fileMenu.addAction(id_save_as)

        exitAction = QAction( '&Выход', self) #QIcon('exit.png'),
        exitAction.setShortcut('Esc')
        exitAction.setStatusTip('Выход и программы')
        exitAction.triggered.connect(lambda:self.closeEvent(QCloseEvent()))
        fileMenu.addAction(exitAction)
      
    
        
        editMenu = menubar.addMenu('&Правка')


        add_row = QAction( '&Добавить ветвь', self) #QIcon('exit.png'),
        add_row.setShortcut('Ctrl+Q')
        add_row.setStatusTip('Добавить ветвь в таблицу')
        add_row.triggered.connect(self.AddBranch)
        editMenu.addAction(add_row)

        remove_row = QAction( '&Удалить ветвь', self) #QIcon('exit.png'),
        remove_row.setShortcut('Ctrl+A')
        remove_row.setStatusTip('Удалить ветвь из таблицы')
        remove_row.triggered.connect(self.RemoveBranch) 
        editMenu.addAction(remove_row)

        add_sector = QAction( '&Добавить участок', self) #QIcon('exit.png'),
        add_sector.setShortcut('Ctrl+W')
        add_sector.setStatusTip('Добавить участок в таблицу')
        add_sector.triggered.connect(self.AddSector)
        editMenu.addAction(add_sector)

        remove_sector = QAction( '&Удалить участок', self) #QIcon('exit.png'),
        remove_sector.setShortcut('Ctrl+S')
        remove_sector.setStatusTip('Удалить участок из таблицы')
        remove_sector.triggered.connect(self.RemoveSector)
        editMenu.addAction(remove_sector)

        add_params = QAction( '&Добавить параметры', self) #QIcon('exit.png'),
        add_params.setShortcut('Ctrl+E')
        add_params.setStatusTip('Добавить параметры в таблицу')
        add_params.triggered.connect(self.AddParams)
        editMenu.addAction(add_params)

        remove_params = QAction( '&Удалить параметры', self) #QIcon('exit.png'),
        remove_params.setShortcut('Ctrl+D')
        remove_params.setStatusTip('Удалить параметры из таблицы')
        remove_params.triggered.connect(self.RemoveParams)
        editMenu.addAction(remove_params)

        add_tab = QAction( '&Добавить вкладку', self) #QIcon('exit.png'),
        add_tab.setShortcut("Ctrl+F")
        add_tab.setStatusTip('Добаваить новую вкладку')
        add_tab.triggered.connect(self.AddTab)
        editMenu.addAction(add_tab)

        remove_tab = QAction( '&Удалить вкладку', self) #QIcon('exit.png'),
        remove_tab.setShortcut("Ctrl+G")
        remove_tab.setStatusTip('Удалить текущую вкладку')
        remove_tab.triggered.connect(self.RemoveTab)
        editMenu.addAction(remove_tab)

        clear_tab = QAction( '&Очистить вкладку', self) #QIcon('exit.png'),
        clear_tab.setShortcut("Ctrl+Shift+D")
        clear_tab.setStatusTip('Очистить текущую вкладку')
        clear_tab.triggered.connect(self.ClearTab)
        editMenu.addAction(clear_tab)


        calcMenu = menubar.addMenu('&Расчёт')

        run_calc = QAction( '&Расчёт ОКГТ', self) #QIcon('exit.png'),
        run_calc.setShortcut("Ctrl+R")
        run_calc.setStatusTip('Запустить расчёт ОКГТ')
        run_calc.triggered.connect(self.RunCalculation)
        calcMenu.addAction(run_calc)

        save_memorandum = QAction( '&Создать служебную записку', self) #QIcon('exit.png'),
        #save_memorandum.setShortcut("Ctrl+R")
        save_memorandum.setStatusTip('Создать служебную записку')
        save_memorandum.triggered.connect(self.CreateMemorandumDoc)
        calcMenu.addAction(save_memorandum)

        save_explanatory = QAction( '&Создать пояснительную записку', self) #QIcon('exit.png'),
        #save_memorandum.setShortcut("Ctrl+R")
        save_explanatory.setStatusTip('Создать пояснительнуюю записку')
        save_explanatory.triggered.connect(self.CreateExplanatoryDoc)
        calcMenu.addAction(save_explanatory)

        get_excel_data = QAction( '&Загрузить токи КЗ из Excel', self) #QIcon('exit.png'),
        #save_memorandum.setShortcut("Ctrl+R")
        get_excel_data.setStatusTip('Загрузить токи КЗ из Excel')
        get_excel_data.triggered.connect(self.getExcelIscData)
        calcMenu.addAction(get_excel_data)


        settingsMenu = menubar.addMenu('&Настройки')

        save_recepients_settings = QAction( '&Сохранить параметры отчёта', self) #QIcon('exit.png'),
        #run_calc.setShortcut("Ctrl+R")
        save_recepients_settings.setStatusTip('Сохранить параметры отчёта по умолчанию')
        save_recepients_settings.triggered.connect(self.SaveReportSettings)
        settingsMenu.addAction(save_recepients_settings)



    def AddBranch(self):
        ind = self.mainTabWidget.currentIndex()
        if ind == 0:
            self.Okgt_node_table.add_row()
        elif ind == 1:
            self.Ps_table.add_row()
        elif ind == 2:
            ind = self.Vl_Tabs.currentIndex()
            if ind>-1:
                wd = self.Vl_Tabs.widget(ind)
                data = self.vl_liks[wd]
                data["branches"].add_row()
        elif ind == 3:
            ind = self.Rpa_Tabs.currentIndex()
            if ind>-1:
                wd = self.Rpa_Tabs.widget(ind)
                data = self.rpa_liks[wd]
                data['rpa_settings'].add_row()
        

    def AddSector(self):
        ind = self.mainTabWidget.currentIndex()
        if ind == 0:
            self.Okgt_sector_table.add_row()
        elif ind == 2:
            ind = self.Vl_Tabs.currentIndex()
            if ind>-1:
                wd = self.Vl_Tabs.widget(ind)
                data = self.vl_liks[wd]
                data["sector"].add_row()
        elif ind == 3:
            ind = self.Rpa_Tabs.currentIndex()
            if ind>-1:
                wd = self.Rpa_Tabs.widget(ind)
                data = self.rpa_liks[wd]
                data['sc_table'].add_row()
        

    def AddParams(self):
        ind = self.mainTabWidget.currentIndex()
        if ind == 0:
            self.Okgt_single_table.add_row()
        elif ind == 2:
            ind = self.Vl_Tabs.currentIndex()
            if ind>-1:
                wd = self.Vl_Tabs.widget(ind)
                data = self.vl_liks[wd]
                p_ind = data["params_tab"].currentIndex()
                header_t = data["params_tab"].tabText(p_ind)
                data["params"][self.vl_settings_dict_rev[header_t]].add_row()
        

    def RemoveBranch(self):
        ind = self.mainTabWidget.currentIndex()
        if ind == 0:
            self.Okgt_node_table.remove_row()
        elif ind == 1:
            self.Ps_table.remove_row()
        elif ind == 2:
            ind = self.Vl_Tabs.currentIndex()
            if ind>-1:
                wd = self.Vl_Tabs.widget(ind)
                data = self.vl_liks[wd]
                data["branches"].remove_row()

        elif ind == 3:
            ind = self.Rpa_Tabs.currentIndex()
            if ind>-1:
                wd = self.Rpa_Tabs.widget(ind)
                data = self.rpa_liks[wd]
                data['rpa_settings'].remove_row()
        

    def RemoveSector(self):
        ind = self.mainTabWidget.currentIndex()
        if ind == 0:
            self.Okgt_sector_table.remove_row()
        elif ind == 2:
            ind = self.Vl_Tabs.currentIndex()
            if ind>-1:
                wd = self.Vl_Tabs.widget(ind)
                data = self.vl_liks[wd]
                data["sector"].remove_row()
        elif ind == 3:
            ind = self.Rpa_Tabs.currentIndex()
            if ind>-1:
                wd = self.Rpa_Tabs.widget(ind)
                data = self.rpa_liks[wd]
                data['sc_table'].removed_row()
        

    def RemoveParams(self):
        ind = self.mainTabWidget.currentIndex()
        if ind == 0:
            self.Okgt_single_table.remove_row()
        elif ind == 2:
            ind = self.Vl_Tabs.currentIndex()
            if ind>-1:
                wd = self.Vl_Tabs.widget(ind)
                data = self.vl_liks[wd]
                p_ind = data["params_tab"].currentIndex()
                header_t = data["params_tab"].tabText(p_ind)
                data["params"][self.vl_settings_dict_rev[header_t]].remove_row()


    def ShowMessage(self, text):
        Message = QMessageBox(QMessageBox.Question,  f'Удаление вкладки {text}',
                    f"Вы дейстивлеьно хотите удалить вкладку {text}?", parent=self)
        Message.addButton('Да', QMessageBox.YesRole)
        Message.addButton('Нет', QMessageBox.NoRole)
        #Message.addButton('Сохранить', QMessageBox.ActionRole)
        reply = Message.exec()
        return reply == 0

        

    def ClearTab(self):
        ind = self.mainTabWidget.currentIndex()
        if ind == 0:
            if self.ShowMessage(self.mainTabWidget.tabText(ind)):
                self.Okgt_node_table.clear_table()
                self.Okgt_sector_table.clear_table()
                self.Okgt_single_table.clear_table()
            
        elif ind == 1:
            if self.ShowMessage(self.mainTabWidget.tabText(ind)):
                self.Ps_table.clear_table()
        elif ind == 2:
            ind = self.Vl_Tabs.currentIndex()
            if ind>-1:
                if self.ShowMessage(self.mainTabWidget.tabText(ind)):
                    wd = self.Vl_Tabs.widget(ind)
                    data = self.vl_liks[wd]
                    data["branches"].clear_table()
                    data["sector"].clear_table()
                    for table in data["params"].values():
                        table.clear_table()
        elif ind == 3:
            ind = self.Rpa_Tabs.currentIndex()
            if ind>-1:
                if self.ShowMessage(self.mainTabWidget.tabText(ind)):
                    wd = self.Rpa_Tabs.widget(ind)
                    data = self.rpa_liks[wd]
                    data['rpa_settings'].clear_table()
                    data['sc_table'].clear_table()


    def AddTab(self):
        ind = self.mainTabWidget.currentIndex()
        if ind == 2:
            print("Add VL")
            self.add_vl_tab()
        elif ind == 3:
            print("Add RPA")
            self.add_rpa_tab()

    def RemoveTab(self):
        ind = self.mainTabWidget.currentIndex()
        if ind == 2:
            self.remove_vl_tab()
            print("Remove VL")
        elif ind == 3:
            self.remove_rpa_tab()
            print("Remove RPA")

    
    def getExcelIscData(self):
        try:
            ind = self.mainTabWidget.currentIndex()
            if ind == 3:
                ind = self.Rpa_Tabs.currentIndex()
                if ind>-1:
                    fname = QFileDialog.getOpenFileName(self, 'Открыть файл Excel', self.main_settings['excel_open'],"*.xlsx")
                    if fname[0] == "" and  fname[1] == "": return
                    fname = fname[0]
                    
                    wb = load_workbook(filename = fname)
                    
                    sheets = wb.sheetnames
                    
                    Message = CustomDialog(list(sheets),self)
                    reply = Message.exec()
                    
                    if reply != 0:
                        Kat = wb[Message.cb.currentText()]
                        j = 1
                        I_sc, L_sc = [], []
                        while type(Kat.cell(row=j,column=1).value) in (int,float) and type(Kat.cell(row=j,column=reply).value) in (int,float):
                            L_sc.append(round(Kat.cell(row=j,column=1).value,3))
                            I_sc.append(round(Kat.cell(row=j,column=reply).value,3))
                            j+=1

                        if I_sc and L_sc:
                            wd = self.Rpa_Tabs.widget(ind)
                            data = self.rpa_liks[wd]
                            data['sc_table'].clear_table()
                            data['sc_table'].write_table(1,{1:{"I_sc":I_sc,"L_sc":L_sc}})

                            self.refreshFigure(wd,{"I_sc":I_sc,"L_sc":L_sc})
                        else:
                            raise Exception("Выбранные столбцы не содержат цифр")


        except Exception as ex:
            ems = QErrorMessage(self)
            ems.setWindowTitle('Возникла ошибка')
            ems.showMessage(f'Не получилось прочитать файл Excel ({str(ex)})')
        else:
            QMessageBox.information(self, 'Чтение Excel','Операция прошла успешно.',
                                          buttons=QMessageBox.Ok,
                                          defaultButton=QMessageBox.Ok)

        
                
        


    def okgt_tab_maker(self):
        self.Okgt_node_table = NodeTable()
        self.Okgt_sector_table = OkgtSectorTable(self.Okgt_node_table,self.cntr_pr)
        self.Okgt_single_table = OkgtSingleTable(self.Okgt_sector_table,self.cntr_pr)

        self.Okgt_sector_table.setMinimumHeight(200)

        
        okgt_spltV = QSplitter(Qt.Vertical)
        okgt_spltV.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        okgt_spltV.addWidget(self.Okgt_node_table)
        okgt_spltV.addWidget(self.Okgt_sector_table)
        okgt_spltV.addWidget(self.Okgt_single_table)
        okgt_spltV.setStretchFactor(1, 3) 
        okgt_spltV.setStretchFactor(3, 1) 

        
        okgt_vbl = QVBoxLayout()
        okgt_vbl.addWidget(okgt_spltV)
        self.Okgt_Widget.setLayout(okgt_vbl)

    def ps_tab_maker(self):
        self.Ps_table = PSTable()
        ps_vbl = QVBoxLayout()
        ps_vbl.addWidget(self.Ps_table)
        self.Ps_Widget.setLayout(ps_vbl)


    def add_vl_tab(self):
        line = QLineEdit()
        
        vl_branches = NodeTable()
        
        vl_sectors = VlSectorTable(vl_branches,self.Okgt_sector_table,self.cntr_pr)
        vl_sectors.setMinimumHeight(180)
        
        d = {}
        param_tabs = QTabWidget()
        for k,v in self.vl_settings_dict.items():
            if k == "PSs":
                table = VlPsParamsTable(vl_branches,self.Ps_table,self.cntr_pr)
            elif k == "commonchains":
                table = VlCommonChainsTable(vl_branches,self.vl_liks,self.le_manager,line,self.cntr_pr) 
            else:
                table = VlParamsTable(vl_branches,k,self.cntr_pr)
                
            param_tabs.addTab(table,v)
            d[k] = table
        
        vl_spltV = QSplitter(Qt.Vertical)
        vl_spltV.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        v_top = QVBoxLayout()
        h_top = QHBoxLayout()
        h_top.addWidget(QLabel("Название ВЛ:"))
        h_top.addWidget(line)
        v_top.addLayout(h_top)
        v_top.addWidget(vl_branches)
        v_top.setContentsMargins(0, 0, 0, 0)
        top = QWidget()
        top.setLayout(v_top)

        vl_spltV.addWidget(top)
        vl_spltV.addWidget(vl_sectors)
        vl_spltV.addWidget(param_tabs)
        vl_spltV.setStretchFactor(1, 3) 
        vl_spltV.setStretchFactor(3, 2) 
        vl_spltV.setContentsMargins(11, 5, 11, 0)

        ind = self.Vl_Tabs.addTab(vl_spltV,"")

        self.le_manager.add_lineEdit(line,vl_spltV)
        name = self.le_manager.get_unique_name("ВЛ № ")
        line.setText(name)
        self.Vl_Tabs.setCurrentIndex(ind)

        self.vl_liks[vl_spltV] = {
            "line":line,
            "branches":vl_branches,
            "sector":vl_sectors,
            "params_tab":param_tabs,
            "params":d,
        }


    def remove_vl_tab(self, s_int=None):
        ind = self.Vl_Tabs.currentIndex() if s_int is None else s_int
        if ind>-1:
            if s_int is None:
                Message = QMessageBox(QMessageBox.Question,  'Удаление вкладки ВЛ',
                    f"Вы дейстивлеьно хотите удалить вкладку {self.Vl_Tabs.tabText(ind)}?", parent=self)
                Message.addButton('Да', QMessageBox.YesRole)
                Message.addButton('Нет', QMessageBox.NoRole)
                #Message.addButton('Сохранить', QMessageBox.ActionRole)
                reply = Message.exec()
                if reply == 1:
                    return  

            wd = self.Vl_Tabs.widget(ind)
            data = self.vl_liks[wd]

            data["params"]["commonchains"].clear_table()
            for dtt in self.vl_liks.values():
                dtt["params"]["commonchains"].vlComboEventResolution = False
            
            self.le_manager.remove_lineEdit(data["line"])
            self.Vl_Tabs.removeTab(ind)
            
            del self.vl_liks[wd]

            for dtt in self.vl_liks.values():
                dtt["params"]["commonchains"].vlComboEventResolution = True
            

    def add_rpa_tab(self):
        vl_combo = UserComboBox(self.cntr_pr)
        self.le_manager.add_child(vl_combo)
        

        ps_combo = UserComboBox(self.cntr_pr)
        self.Ps_table.add_child(((0,),),ps_combo)

        check = QCheckBox('')
        check.setCheckState(Qt.Checked)
        
        t_auto = QDoubleSpinBox()
        t_auto.setMinimum(0)
        t_auto.setDecimals(3)
        t_auto.setSingleStep(0.01)
        t_auto.setValue(0.02)

        t_switch = QDoubleSpinBox()
        t_switch.setMinimum(0)
        t_switch.setDecimals(3)
        t_switch.setSingleStep(0.01)
        t_switch.setValue(0.13)
        
        arc_times = QSpinBox()
        arc_times.setMinimum(0)
        arc_times.setSingleStep(1)
        arc_times.setValue(0)
        #arc_times.editingFinished.connect()

        rpa_settings = RPASettingsTable()
        check.stateChanged.connect(lambda x: rpa_settings.setRelativeState(True if x==Qt.Checked else False))
        arc_times.valueChanged.connect(rpa_settings.setColumn)

        sc_table = ShortCircuitLineTable()
        sc_table.setMaximumWidth(240)
        

        settingsGrid = QGridLayout()
        settingsGrid.addWidget(QLabel('ВЛ:'), 0,0)
        settingsGrid.addWidget(QLabel('ПС:'), 1,0)
        settingsGrid.addWidget(check, 2,0)

        settingsGrid.addWidget(vl_combo, 0,1)
        settingsGrid.addWidget(ps_combo, 1,1)
        settingsGrid.addWidget(QLabel('Относительно времени уставки'), 2,1)

        settingsGrid.addWidget(QLabel('Время автоматики:'), 0,2)
        settingsGrid.addWidget(QLabel('Время выключателя:'), 1,2)
        settingsGrid.addWidget(QLabel('Работа АПВ, раз:'), 2,2)

        settingsGrid.addWidget(t_auto, 0,3)
        settingsGrid.addWidget(t_switch, 1,3)
        settingsGrid.addWidget(arc_times, 2,3)

        settingsGrid.addItem(QSpacerItem(100,0, QSizePolicy.Expanding, QSizePolicy.Fixed),0,4)
        settingsGrid.addItem(QSpacerItem(100,0, QSizePolicy.Expanding, QSizePolicy.Fixed),1,4)
        settingsGrid.addItem(QSpacerItem(100,0, QSizePolicy.Expanding, QSizePolicy.Fixed),2,4)


        Vlayout = QVBoxLayout()
        Vlayout.addLayout(settingsGrid)
        Vlayout.addWidget(rpa_settings)

        top_widget = QWidget()
        top_widget.setLayout(Vlayout)

        fg = plt.figure(dpi=75) # Создаём фигуру графика 
        fg_widget = FigureCanvas(fg) # Помещаем фигуру в контейнер
        ax = fg.add_subplot(111) #
        ax.set_xlabel('L, км',fontsize=self.font_s) 
        ax.set_ylabel('Iкз, кА',fontsize=self.font_s)
        ax.set_title("Кривая тока КЗ",fontsize=self.font_s)
        ax.tick_params(labelsize=self.font_s)
        ax.grid(True)
        
        

        rpa_spltV = QSplitter(Qt.Vertical)
        rpa_spltV.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        rpa_spltV.addWidget(top_widget)
        rpa_spltV.addWidget(fg_widget)
        rpa_spltV.setStretchFactor(1, 2) 
        
        rpa_spltH = QSplitter(Qt.Horizontal)
        rpa_spltH.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        rpa_spltH.addWidget(rpa_spltV)
        rpa_spltH.addWidget(sc_table)
        rpa_spltH.setStretchFactor(6, 1)
        
        ind = self.Rpa_Tabs.addTab(rpa_spltH, f"РЗА№{self.Rpa_Tabs.count()+1}")
        self.Rpa_Tabs.setCurrentIndex(ind)

        vl_combo.currentTextChanged.connect(lambda t, w=rpa_spltH:self.renameRPATab(t,w))
        ps_combo.currentTextChanged.connect(lambda t, w=rpa_spltH:self.renameRPATab(t,w))
        sc_table.setPlotDataFunc(lambda d,w=rpa_spltH:self.refreshFigure(w,d))
        

        self.rpa_liks[rpa_spltH] = {
            'vl_combo':vl_combo,
            'ps_combo':ps_combo,
            't_switch':t_switch,
            't_auto':t_auto,
            'arc_times':arc_times,
            'check':check,
            'rpa_settings':rpa_settings,
            'sc_table':sc_table,
            'figure':fg,
            'axis':ax,
            'fg_widget':fg_widget,
        }

        self.activateWindow()

    def remove_rpa_tab(self, s_int=None):
        ind = self.Rpa_Tabs.currentIndex() if s_int is None else s_int
        if ind>-1:
            if s_int is None:
                Message = QMessageBox(QMessageBox.Question,  'Удаление вкладки РЗА',
                    f"Вы дейстивлеьно хотите удалить вкладку {self.Rpa_Tabs.tabText(ind)}?", parent=self)
                Message.addButton('Да', QMessageBox.YesRole)
                Message.addButton('Нет', QMessageBox.NoRole)
                #Message.addButton('Сохранить', QMessageBox.ActionRole)
                reply = Message.exec()
                if reply == 1:
                    return  
                
            wd = self.Rpa_Tabs.widget(ind)
            data = self.rpa_liks[wd]
            self.le_manager.remove_child(data['vl_combo'])
            self.Ps_table.remove_child(((0,),),data['ps_combo'])
            data['vl_combo'].currentTextChanged.disconnect()
            data['ps_combo'].currentTextChanged.disconnect()
            data['check'].stateChanged.disconnect()
            data['arc_times'].valueChanged.disconnect()
            plt.close(data['figure'])

            self.Rpa_Tabs.removeTab(ind)
            del self.rpa_liks[wd]



    def refreshFigure(self,widget,data):
        #I_sc, L_sc =  data['I_sc'], data['L_sc']
        I_sc, L_sc, borders = I_sc_corector(data['I_sc'], data['L_sc'], borders=True)
        ax = self.rpa_liks[widget]['axis']
        title = self.rpa_liks[widget]['ps_combo'].currentText()+' - '+self.rpa_liks[widget]['vl_combo'].currentText()
        ax.clear()
        ax.set_xlabel('L, км',fontsize=self.font_s) 
        ax.set_ylabel('Iкз, кА',fontsize=self.font_s)
        ax.set_title(title,fontsize=self.font_s, loc='left') #"Кривая тока КЗ"
        ax.tick_params(labelsize=self.font_s)
        ax.grid(True) 
        ax.plot(L_sc,I_sc,'r')#,label=self.FAName
        #fig.suptitle('This figure suptitle should be on the left', horizontalalignment = 'left')
        
        ax.set_xlim(borders)
        
        xloc = plt.MaxNLocator()
        ax.xaxis.set_major_locator(xloc)
        self.rpa_liks[widget]['fg_widget'].draw()
        

    def renameRPATab(self,_,widget):
        ind = self.Rpa_Tabs.indexOf(widget)
        vl_combo = self.rpa_liks[widget]['vl_combo']
        ps_combo = self.rpa_liks[widget]['ps_combo']
        self.Rpa_Tabs.setTabText(ind,vl_combo.currentText()+"-"+ps_combo.currentText())

    
    def RunCalculation(self):
        
        okgt_info_new, ps_info_new, vl_info_new,  rpa_info_new = self.ReadTables()
        self.okgt_calc_tread.setData(okgt_info_new, vl_info_new, ps_info_new, rpa_info_new)

        if not self.okgt_calc_tread.isRunning():
            self.OkgtCalculationSignals("init:")
            self.okgt_calc_tread.start()
        
    @traceback_erors
    def OkgtCalculationSignals(self,signal):
        if signal == "init:":
            self.okgt_proces_dialog = QProgressDialog("Иницализация расчёта",'Отмена', 0, 0, self)
            self.okgt_proces_dialog.setWindowTitle('Расчет ОКГТ')
            self.okgt_proces_dialog.setMinimumDuration(0)
            self.okgt_proces_dialog.setWindowModality(Qt.WindowModal)
            self.okgt_proces_dialog.canceled.connect(self.okgt_calc_tread.stop)
            self.okgt_proces_dialog.show()
        elif signal == "start:":
            self.okgt_proces_dialog.setLabelText("Запуск расчёта")
        elif signal.startswith("run_messages:"):
            self.okgt_proces_dialog.setLabelText(signal[len("run_messages:"):]) 
        elif signal.startswith("run_points:"):
            data = json.loads(signal[len("run_points:"):])
            mes = f"ВЛ: {data['vl_name']}, ветвь: {data['branch']}, опора: {data['support']}, {data['current']} из {data['length']}"
            self.okgt_proces_dialog.setLabelText(mes) 
            if self.okgt_proces_dialog.maximum() != int(data['length']):
                self.okgt_proces_dialog.setMaximum(int(data['length']))
            self.okgt_proces_dialog.setValue(int(data['current']))
        elif signal.startswith("error:"): 
            self.okgt_proces_dialog.close()
            ems = QErrorMessage(self)
            self.okgt_proces_dialog.canceled.disconnect()
            self.okgt_proces_dialog = None
            ems.setWindowTitle('Возникла ошибка')
            ems.showMessage('В исходных данных допущена ошибка. Проверте исходные данные. ('+signal[len("error:"):]+')')
        elif signal == "end:":
            self.okgt_proces_dialog.close()
            self.okgt_proces_dialog.canceled.disconnect()
            self.okgt_proces_dialog = None
            print("calc is end")
            self.setResultsPlots()

    def setResultsPlots(self):
        for ind in range(self.Rez_Tabs.count()-1,-1,-1):
            wd = self.Rez_Tabs.widget(ind)
            if wd in self.resFig:
                plt.close(self.resFig[wd][0])
                del self.resFig[wd]
            self.Rez_Tabs.removeTab(ind)
        self.sectorsFig = {}
        self.calc_results = self.okgt_calc_tread.getData()
        for val in self.calc_results.values():
            for sector in val["sectors"]:
                if sector[1] != 'single_dielectric':
                    st, ed = sector[2:4]

                    fg = plt.figure(dpi=75) # Создаём фигуру графика 
                    fg_widget = FigureCanvas(fg) # Помещаем фигуру в контейнер
                    ax = fg.add_subplot(111) #
                    ax.plot(val["L"][st:ed],val["B"][st:ed],'r',label="Расчетный тепловой импульс")
                    ax.plot(val["L"][st:ed],val["Bmax"][st:ed],'b',label="Допустимый тепловой импульс")
                    ax.set_xlabel('Растояние, км',fontsize=self.font_s) 
                    ax.set_ylabel('Тепловой импульс, кА\u00B2·c',fontsize=self.font_s)
                    ax.set_title(f"{sector[0]}",fontsize=self.font_s)
                    ax.tick_params(labelsize=self.font_s)
                    ax.legend(frameon=False,fontsize=self.font_s) # Выводим легенду графика
                    ax.set_xlim([val["L"][st],val["L"][ed-1]])
                    ax.grid(True)

                    self.Rez_Tabs.addTab(fg_widget, f"{sector[0]}")

                    if sector[4]:
                        self.Rez_Tabs.tabBar().setTabTextColor(self.Rez_Tabs.indexOf(fg_widget),QColor(255,0,0))
                    

                    self.resFig[fg_widget] = (fg,ax)
                    self.sectorsFig[sector] = (fg,ax)

        self.mainTabWidget.setCurrentIndex(4)

    #@traceback_erors 
    def ToJsonFormat(self,okgt_info,ps_info,vl_info,rpa_info):
        def ReadData(file):
            if type(file) == dict:
                new_file = {}
                for key, val in file.items():
                    nkey = self.jsSpleater.join(key) if type(key) == tuple else key
                    if type(val) == tuple:
                        nval = self.jsSpleater.join(val)
                    elif type(val) == dict or type(val) == list:
                        nval = ReadData(val)
                    else:
                        nval = val
                    new_file[nkey] = nval
                return new_file

            elif type(file) == list:
                new_file = []
                for val in file:
                    if type(val) == tuple:
                        nval = self.jsSpleater.join(val)
                    elif type(val) == dict or type(val) == list:
                        nval = ReadData(val)
                    else:
                        nval = val
                    new_file.append(nval)
                return new_file

        okgt_info_new = ReadData(okgt_info)
        ps_info_new = ReadData(ps_info)
        vl_info_new = ReadData(vl_info)
        rpa_info_new = ReadData(rpa_info)
        

        return okgt_info_new,ps_info_new,vl_info_new,rpa_info_new


    #@traceback_erors
    def FromJsonFormat(self,okgt_info,ps_info,vl_info,rpa_info):
        def ReadData(file):
            if type(file) == dict:
                new_file = {}
                for key, val in file.items(): 
                    nkey = tuple(key.split(self.jsSpleater)) if key.find(self.jsSpleater)!=-1 else key
                    if type(val) == str:
                        nval = tuple(val.split(self.jsSpleater)) if val.find(self.jsSpleater)!=-1 else val
                    elif type(val) == dict or type(val) == list:
                        nval = ReadData(val)
                    else:
                        nval = val
                    new_file[nkey] = nval
                return new_file

            elif type(file) == list:
                new_file = []
                for val in file: 
                    if type(val) == str:
                        nval = tuple(val.split(self.jsSpleater)) if val.find(self.jsSpleater)!=-1 else val
                    elif type(val) == dict or type(val) == list:
                        nval = ReadData(val)
                    else:
                        nval = val
                    new_file.append(nval)
                return new_file

        okgt_info_new = ReadData(okgt_info)
        ps_info_new = ReadData(ps_info)
        vl_info_new = ReadData(vl_info)
        rpa_info_new = ReadData(rpa_info)

        return okgt_info_new,ps_info_new,vl_info_new,rpa_info_new
        
        
    #@traceback_erors 
    def WriteTables(self,okgt_info,ps_info,vl_info,rpa_info):
        self.Okgt_node_table.write_table(okgt_info)
        self.Okgt_sector_table.write_table(okgt_info)
        self.Okgt_single_table.write_table(okgt_info)
        self.Ps_table.write_table(ps_info)

        for ind in range(self.Vl_Tabs.count()-1,-1,-1):
            self.remove_vl_tab(ind)

        for ind, (name,info) in enumerate(vl_info.items()):
            self.add_vl_tab()
            data = self.vl_liks[self.Vl_Tabs.widget(ind)]
            data['line'].setText(name)
            data["branches"].write_table(info,vl=True)
            data["sector"].write_table(info)
            
            for t_n, table in data["params"].items():
                if t_n!="commonchains":
                    table.write_table(info)

        for ind, info in enumerate(vl_info.values()):
            data = self.vl_liks[self.Vl_Tabs.widget(ind)]
            data["params"]["commonchains"].write_table(info)
            

        for ind in range(self.Rpa_Tabs.count()-1,-1,-1):
            self.remove_rpa_tab(ind)

        for ind, ((vl_name,ps_name),info) in enumerate(rpa_info.items()):
            self.add_rpa_tab()
            data = self.rpa_liks[self.Rpa_Tabs.widget(ind)]
            data['vl_combo'].setCurrentText(vl_name)
            data['ps_combo'].setCurrentText(ps_name)
            data['t_switch'].setValue(info["Tswitch"])
            data['t_auto'].setValue(info["Tautomation"])
            data['arc_times'].setValue(info["arc_times"])
            data['rpa_settings'].write_table((vl_name,ps_name),rpa_info)
            data['sc_table'].write_table((vl_name,ps_name),rpa_info)

            self.refreshFigure(self.Rpa_Tabs.widget(ind),data['sc_table'].read_table(own=True))

        self.activateWindow()

        
    #@traceback_erors
    def ReadTables(self):
        lst = self.Okgt_node_table.read_table()
        dct = self.Okgt_sector_table.read_table(lst)
        okgt_info_new = self.Okgt_single_table.read_table(dct)
        ps_info_new = self.Ps_table.read_table()

        d_lst = {}
        d_com_tables = {}
        vl_info_new = {}
        for data in self.vl_liks.values():
            name = data['line'].text().strip()
            br_lst = data["branches"].read_table()
            d_lst[name] = br_lst
            d_com_tables[name] = data["params"]["commonchains"]
            vl_info_new[name] = {}

            vl_info_new[name].update([(key,val) for key,val in data["sector"].read_table(br_lst).items()])
            for t_n, table in data["params"].items():
                if t_n!="commonchains":
                    vl_info_new[name].update([(key,val) for key,val in table.read_table(br_lst).items()])

        
        for name, table in d_com_tables.items():
            vl_info_new[name].update([(key,val) for key,val in table.read_table(name,d_lst).items()])

        #vl_info = vl_info_new

    
        rpa_info_new = {}
        for data in self.rpa_liks.values():
            vl_name = data['vl_combo'].currentText()
            ps_name = data['ps_combo'].currentText()

            if vl_name!='Нет' and ps_name!='Нет':
                d1 = data['rpa_settings'].read_table()
                d2 = data['sc_table'].read_table()
                
                rpa_info_new[(vl_name,ps_name)] ={
                    "Tswitch":data['t_switch'].value(),
                    "Tautomation":data['t_auto'].value(),
                    "arc_times":data['arc_times'].value(),
                    **d1,
                    **d2,
                }

        
        return okgt_info_new, ps_info_new, vl_info_new, rpa_info_new

    def InitialDataOpen(self):
        try:
            fname = QFileDialog.getOpenFileName(self, 'Открыть файл ИД', self.main_settings['id_open'],'*.okgt')
            if fname[0] == "" and  fname[1] == "": return
            fname = fname[0]

            self.main_settings['id_open'] = os.path.dirname(fname)
            self.currentFilePath = fname
            self.currentFileName = os.path.splitext(os.path.split(fname)[1])[0]
            self.setWindowTitle(f"OKGT - {self.currentFileName}")

            with open(fname, "r", encoding="utf8") as f:
                data  = json.load(f)
            
            
            self.WriteTables(*self.FromJsonFormat(**{i:data[i] for i in ['okgt_info','ps_info','vl_info','rpa_info']}))
            if "report_settings" in data:
                self.setReportSettings(data["report_settings"])
            else:
                self.setReportSettings(self.main_settings)
            

        except Exception as ex:
            ems = QErrorMessage(self)
            ems.setWindowTitle('Возникла ошибка')
            ems.showMessage('Не получилось открыть файл. '+
                            'Вероятнее всего файл поврежден'+str(ex))

    def InitialDataSave(self):
        try:
            if os.path.exists(self.currentFilePath):
                Message = QMessageBox(QMessageBox.Question,  'Сохранить',
                    "Вы дейстивлеьно хотите заменить "+self.currentFileName+"?", parent=self)
                Message.addButton('Да', QMessageBox.YesRole)
                Message.addButton('Нет', QMessageBox.NoRole)
                reply = Message.exec()
                if reply == 0:
                    fname = self.currentFilePath
                    okgt_info, ps_info, vl_info, rpa_info = self.ToJsonFormat(*self.ReadTables())

                    data = {
                        'okgt_info':okgt_info,
                        'ps_info':ps_info,
                        'vl_info':vl_info,
                        'rpa_info':rpa_info,
                        "report_settings":self.getReportSettings(),
                    }

                    with open( fname, "w", encoding="utf8") as write_file:
                            json.dump(data, write_file, indent=4)

        except Exception as ex:
            ems = QErrorMessage(self)
            ems.setWindowTitle('Возникла ошибка')
            ems.showMessage('Не получилось сохранить файл. '+
                            'Проверьте введённые данные а также сохраняете ли вы уже в открытый файл.'+str(ex))
        

    #@traceback_erors
    def InitialDataSaveAs(self):
        try:
            fname = QFileDialog.getSaveFileName(self, 'Сохранить файл ИД как', os.path.join(self.main_settings['id_save_as'],self.currentFileName),'*.okgt')
            if fname[0] == "" and fname[1] == "": return
            fname = fname[0]
            self.main_settings['id_save_as'] = os.path.dirname(fname)
            self.currentFilePath = fname
            self.currentFileName = os.path.splitext(os.path.split(fname)[1])[0]
            self.setWindowTitle(f"OKGT - {self.currentFileName}")

            okgt_info, ps_info, vl_info, rpa_info = self.ToJsonFormat(*self.ReadTables())

            data = {
                'okgt_info':okgt_info,
                'ps_info':ps_info,
                'vl_info':vl_info,
                'rpa_info':rpa_info,
                "report_settings":self.getReportSettings(),
            }

            with open( fname, "w", encoding="utf8") as write_file:
                    json.dump(data, write_file, indent=4)


        except Exception as ex:
            ems = QErrorMessage(self)
            ems.setWindowTitle('Возникла ошибка')
            ems.showMessage('Не получилось сохранить файл. '+
                            'Проверьте введённые данные а также сохраняете ли вы уже в открытый файл.'+str(ex))
        else:
            QMessageBox.information(self, 'Сохранение в файл','Операция прошла успешно.',
                                          buttons=QMessageBox.Ok,
                                          defaultButton=QMessageBox.Ok)
        

    def closeEvent(self, event):
        Message = QMessageBox(QMessageBox.Question,  'Выход из программы',
            "Вы дейстивлеьно хотите выйти?", parent=self)
        Message.addButton('Да', QMessageBox.YesRole)
        Message.addButton('Нет', QMessageBox.NoRole)
        #Message.addButton('Сохранить', QMessageBox.ActionRole)
        reply = Message.exec()
        if reply == 0:  
            with open(os.path.join(self.path_midle_files,'main_settings.json'), "w", encoding="utf8") as f:
                json.dump(self.main_settings,f, indent=4)
            qApp.quit()
        elif reply == 1:
            event.ignore()


if __name__=='__main__':
    
    app=QApplication(sys.argv)
    window = MyWindow()
    window.show()
    sys.exit(app.exec_())

    