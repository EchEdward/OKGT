import numpy as np

from scipy.integrate import simps
from scipy.constants import epsilon_0
from scipy.optimize import curve_fit
from scipy.interpolate import interp1d

import scipy.sparse as sparse
import scipy.sparse.linalg as linalg

from openpyxl import load_workbook



def load_data():
    """ Load catalog of parametrs of cunductors and supports  """
    Katal = load_workbook(filename = 'Katal.xlsx')
    Kat = Katal['Kat']

    supports = {}
    conductors = {}

    s = (2,3,4,5,6,9,10,11,12,13,14,14)
    c = (18,19)
    check = lambda arr,j:sum([0 if Kat.cell(row=i,column=j).value is not None else 1 for i in arr])

    j = 2
    while  Kat.cell(row=1,column=j).value is not None or\
        Kat.cell(row=17,column=j).value is not None:
        if not check(s,j):
            supports[Kat.cell(row=1,column=j).value] = {
                "X":{
                    "A":Kat.cell(row=2,column=j).value,
                    "B":Kat.cell(row=3,column=j).value,
                    "C":Kat.cell(row=4,column=j).value,
                    "T1":Kat.cell(row=5,column=j).value,
                    "T2":Kat.cell(row=6,column=j).value,
                },
                "Y":{
                    "A":Kat.cell(row=9,column=j).value,
                    "B":Kat.cell(row=10,column=j).value,
                    "C":Kat.cell(row=11,column=j).value,
                    "T1":Kat.cell(row=12,column=j).value,
                    "T2":Kat.cell(row=13,column=j).value,
                },
                "type":Kat.cell(row=14,column=j).value,
                "isulator":Kat.cell(row=15,column=j).value,
            }
        if not check(c,j):
            conductors[Kat.cell(row=17,column=j).value] = {
                "r":Kat.cell(row=18,column=j).value,
                "R0":Kat.cell(row=19,column=j).value,
                "Id":Kat.cell(row=20,column=j).value,
                "Bsc":Kat.cell(row=21,column=j).value,
                "Grounded_conductor":Kat.cell(row=22,column=j).value,
            }
        j+=1

    return supports, conductors
    
k_supports, k_conductors = load_data()


carson_cashe = {}
mu0=np.pi*4*10**(-7)
jp=1j*2*np.pi*50
sh_int=np.arange(0, 1, 0.00001)

ro_fe =  0.1 #Ом*мм2/м
R_isol = 10**8
R_bypass = 10**-5


def get_vl_sector(key,vl_info):
    lst = {}
    for name, value in vl_info.items():
        #branches = {}
        """ for k, branch in value["branches"].items():
            branches[k] = (branch["supportN"],branch["supportK"]) """
        
        lst[name] = []

        for sector in value["sectors"]:
            if sector["type"] == "with_okgt":
                if sector["link_okgt"] == key:
                    d_sector = {k:v for k,v in sector.items()}
                    #d_sector["branch_supports"] = branches[d_sector["link_branch"]]
                    d_sector["name_vl"] = name
                    lst[name].append(d_sector)

        if len(lst[name])==0:
            del lst[name]
    
    return lst  

def find_entry(dct, length):
    dl = 0.0001
    orders = {}
    for name, lst in dct.items():
        
        ln, lk = 0,  length
        order = []
        len_lst  = len(lst)
        for _ in range(len_lst):
            for i, val in enumerate(lst):
                if i in order: 
                    continue
                if ln-dl<=val['lengthN']<=ln+dl and val['lengthK']<=lk+dl:
                    order.append(i)
                    ln = val['lengthK']

            if len_lst == len(order):
                break
        else:
            
            raise Exception("Entries were not found")
            
        orders[name] = order

    
    return {name:[val[i] for i in orders[name]] for name,val in dct.items()}

def R_cc(r0,lg,dg,h):
    R = r0/(2*np.pi*lg*10**3)*(np.log((2*10**6*lg)/dg)+np.log(lg*10**3/2/h))
    return R

def single_okgt_zy(sector,v):
    grounded = []
    for gr in range(v+1):
        if gr==0 and sector["way_grounded"] in ["left", "both"] and sector["countercable"]:
            grounded.append("both")
        elif gr==0 and sector["way_grounded"] in ["left", "both"] and not sector["countercable"]:
            grounded.append("Yzy")
        elif gr==0 and sector["way_grounded"] not in ["left", "both"] and sector["countercable"]:
            grounded.append("Ycc")
        
        elif gr==v and sector["way_grounded"] in ["right", "both"] and sector["countercable"]:
            grounded.append("both")
        elif gr==v and sector["way_grounded"] in ["right", "both"] and not sector["countercable"]:
            grounded.append("Yzy")
        elif gr==v and sector["way_grounded"] not in ["right", "both"] and sector["countercable"]:
            grounded.append("Ycc")

        elif 0<gr<v and sector["way_grounded"] in ["inside", "both"] and sector["countercable"]:
            grounded.append("both2")
        elif 0<gr<v and sector["way_grounded"] in ["inside", "both"] and not sector["countercable"]:
            grounded.append("Yzy")
        elif 0<gr<v and sector["way_grounded"] not in ["inside", "both"] and sector["countercable"]:
            grounded.append("Ycc2")

        else:
            grounded.append(None)
    
    return grounded

def type_zy_chose(sector,grounded,Y_cc):
    return (1/sector["point_grounded"] if grounded in ["Yzy","both","both2"] else 0)+\
        (Y_cc/2 if grounded in ["Ycc","both"] else 0)+(Y_cc if grounded in ["Ycc2","both2"] else 0)  


def get_vl_sector_info(vl,branch,supports,dtype,vl_info):
    #"phases", "conductors", "groundwires", "supports", "PSs"
    f = lambda ni,ki,ns,ks: (ni<ki and ns>=ni and ks<=ki) or (ni>ki and ns<=ni and ks>=ki)
    if dtype["name"] == "conductors":
        for item in vl_info[vl]["conductors"]:
            if item["link_branch"]==branch and f(item["supportN"],item["supportK"],supports[0],supports[1]):
                return {"c1":item["conductor"],"c2":item["conductor"],"c3":item["conductor"]}

    elif dtype["name"] == "phases":
        for item in vl_info[vl]["phases"]:
            if item["link_branch"]==branch and f(item["supportN"],item["supportK"],supports[0],supports[1]):
                if item["phase"][0] == "-":
                    mirrored = True
                    p1,p2,p3 = item["phase"][1:]
                else:
                    mirrored = False
                    p1,p2,p3 = item["phase"]
                return {"p1":p1,"p2":p2,"p3":p3,"mirrored":mirrored}

    elif dtype["name"] =="supports":
        for item in vl_info[vl]["supports"]:
            if item["link_branch"]==branch and f(item["supportN"],item["supportK"],supports[0],supports[1]):
                return {"s1":item["support"],"s2":item["support"],"s3":item["support"],"s4":item["support"],"s5":item["support"]}

    elif dtype["name"] =="groundwires":
        for item in vl_info[vl]["groundwires"]:
            if item["link_branch"]==branch and f(item["supportN"],item["supportK"],supports[0],supports[1]):
                if item["type"] == "one":
                    if item["is_okgt"] is None:
                        return {"gw1":item["groundwire1"],"gw2":None,"is_okgt":None}
                    elif item["is_okgt"] is not None and item["is_okgt"]=="groundwire1":
                        return {"gw1":item["groundwire1"],"gw2":None, "is_okgt":"gw1"}
                elif item["type"] == "two":
                    if item["is_okgt"] is None:
                        return {"gw1":item["groundwire1"],"gw2":item["groundwire2"],"is_okgt":None}
                    elif item["is_okgt"] is not None and item["is_okgt"]=="groundwire1":
                        return {"gw1":item["groundwire1"],"gw2":item["groundwire2"],"is_okgt":"gw1"}
                    elif item["is_okgt"] is not None and item["is_okgt"]=="groundwire2":
                        return {"gw1":item["groundwire1"],"gw2":item["groundwire2"],"is_okgt":"gw2"}
        else:
            return {"gw1":None,"gw2":None,"is_okgt":None}

    elif dtype["name"] =="PSs":
        for item in vl_info[vl]["PSs"]:
            if vl_info[vl]["branches"][branch][dtype["side"]] == item["PS_name"]:
                return {"length":item["length"]}

    #"countercables"
    elif dtype["name"] =="countercables":
        for item in vl_info[vl]["countercables"]:
            if item["link_branch"]==branch and f(item["supportN"],item["supportK"],supports[0],supports[1]):
                return item
        else:
            return None

    raise Exception("Supports %s don't exist in %s %s %s" % (str(supports),vl,str(branch),dtype["name"]))

def sector_entry(lst,branch):
    new_lst = [i for i,val in enumerate(lst) if val["link_branch"]==branch]
    if len(new_lst) == 0:
        raise Exception("Entries were not found 3")
    count = 0
    for i in new_lst:
        if lst[i]["supportN"]<lst[i]["supportK"]:
            count+=1
    #print(count, len(new_lst))
    if count != len(new_lst) and count != 0:
        raise Exception("Entries were not found 1")

    derection = count == len(new_lst)
    order = sorted(new_lst, key=lambda x: lst[x]["supportN"], reverse=not derection)
    for i in range(1,len(new_lst)):
        if lst[order[i]]["supportN"] != lst[order[i-1]]["supportK"]:
            raise Exception("Entries were not found 2")

    return order

def link_vl_to_vl(vl_info,vl_name,branch,n,k):
    lst = vl_info[vl_name]["commonchains"]
    results = []
    for item in lst:
        Branch, BranchN = item["link_branch"],item["other_link_branch"]
        N,K, Nn,Kn = item["supportN"],item["supportK"],item["other_supportN"],item["other_supportK"]

        if Branch==branch and (N<=n<k<=K or N>=n>k>=K):
            gen = zip(range(N,K+int((K-N)/abs(K-N)),int((K-N)/abs(K-N))),range(Nn,Kn+int((Kn-Nn)/abs(Kn-Nn)),int((Kn-Nn)/abs(Kn-Nn))))
            d = {i:j for i,j in gen}
            results.append((item["other_vl_name"],BranchN,d[n],d[k]))

    return results

def init_carson_data(key,data,trig,shift):
    if key=="phase":
        support = k_supports[data["Tsupport"]]
        mirrored = -1 if data["mirrored"] else 1
        X = mirrored*support["X"][data["phase"]]
        Y = support["Y"][data["phase"]]-support["isulator"]+shift
        R = k_conductors[data["conductor"]]["R0"]
        r = k_conductors[data["conductor"]]["r"]
        return X,Y,R,r
        
    elif key=="groundwire":
        if data["conductor"][1] or (not data["conductor"][1] and data["conductor"][0] is None) or not trig:
            return None,None,R_isol,None
        elif not data["conductor"][1] and data["conductor"][0] is not None and trig:
            support = k_supports[data["Tsupport"]]
            mirrored = -1 if data["mirrored"] else 1
            X = mirrored*support["X"][data["phase"]]
            Y = support["Y"][data["phase"]]
            if Y==0 and X==0:
                raise Exception("%s doesn't support groundwire %s" % (data["Tsupport"],data["phase"]))
            else:
                Y+=shift
            R = k_conductors[data["conductor"][0]]["R0"]
            r = k_conductors[data["conductor"][0]]["r"]
            return X,Y,R,r

    elif key=="okgt_vl":
        support = k_supports[data["Tsupport"]]
        mirrored = -1 if data["mirrored"] else 1
        X = mirrored*support["X"][data["phase"]]
        Y = support["Y"][data["phase"]]
        if Y==0 and X==0:
            raise Exception("%s doesn't support groundwire %s" % (data["Tsupport"],data["phase"]))
        else:
            Y+=shift
        R = k_conductors[data["conductor"][0]]["R0"]
        r = k_conductors[data["conductor"][0]]["r"]
        return X,Y,R,r
    elif key=="cc":
        X = data["X_countercable"]
        Y = data["H_countercable"]+shift
        r = data["D_countercable"]/2*10**-3
        R = (4*ro_fe)/(np.pi*data["D_countercable"]**2)*10**3
        return X,Y,R,r

    elif key=="okgt_single":
        X = data['X_cable']
        Y = data['H_cable']+shift
        R = k_conductors[data['groundwire']]["R0"]
        r = k_conductors[data['groundwire']]["r"]
        return X,Y,R,r

    elif key=="phase_ps":
        return None,None,R_isol,None
    elif key=="groundwire_ps":
        if data["conductor"][0] is None or not trig:
            return None,None,R_isol,None
        elif data["conductor"][0] is not None and trig:
            support = k_supports[data["Tsupport"]]
            mirrored = -1 if data["mirrored"] else 1
            X = mirrored*support["X"][data["phase"]]
            Y = support["Y"][data["phase"]]
            if Y==0 and X==0:
                raise Exception("%s doesn't support groundwire %s" % (data["Tsupport"],data["phase"]))
            else:
                Y+=shift
            R = k_conductors[data["conductor"][0]]["R0"]
            r = k_conductors[data["conductor"][0]]["r"]
            return X,Y,R,r

Cars_ii = lambda x,hi,ri,mp:(np.exp(-2*x*hi)*np.cos(x*ri))/(x+(x**2+mp**2)**0.5)
Cars_ij = lambda x,hi,hj,xi,xj,mp:(np.exp(-x*(hi+hj))*np.cos(x*abs(xi-xj)))/(x+(x**2+mp**2)**0.5)

def Carson(data):
    dd = tuple([tuple(data[i]) if i<5 else data[i] for i in range(7)])
    if dd in carson_cashe:
        return carson_cashe[dd]
    else:
        X,Y,R,r,trig,pz,l = dd 
        mp=(mu0*jp*(1/pz))**0.5
        

        size = len(X)
        M = sparse.lil_matrix((size,size), dtype=np.complex128)

        for a in range(size):
            for b in range(a,size):
                if a == b and trig[a]:
                    M[a,b]=R[a]*l + jp*mu0/2/np.pi*(np.log(2*abs(Y[a])/r[a])+2*simps(Cars_ii(sh_int,Y[a],r[a],mp),sh_int))*l*10**3
                elif a == b and not trig[a]:
                    M[a,b]=R[a]
                elif a!=b and trig[a] and trig[b]:
                    M[a,b]=M[b,a]=jp*mu0/4/np.pi*(np.log(((Y[a]+Y[b])**2+(X[a]-X[b])**2)/((Y[a]-Y[b])**2+(X[a]-X[b])**2))\
                        +4*simps(Cars_ij(sh_int,Y[a],Y[b],X[a],X[b],mp),sh_int))*l*10**3  
                elif a!=b and (trig[a] or trig[b]):
                    M[a,b]=M[b,a]=0


        carson_cashe[dd] = M
        return M
        

def combinations(lst,k):
    l = len(lst)
    if k>l:
        #raise StopIteration
        return
    
    comb = [i for i in range(k)]
    while True:
        yield [lst[i] for i in comb]
        
        if comb[0]==l-k:
            #raise StopIteration
            return
        
        for i in range(k-1,-1,-1):
            comb[i]+=1
            if comb[i]==l-(k-1-i):
                if i-1>-1:
                    for j in range(i-1,-1,-1):
                        a = comb[j]+i-j+1
                        if a<l-(k-1-i):
                            comb[i]=a
                            break
            else:
                break

def submatrix_putter(slices,SM, M):
    for sl in slices:
        M[sl[2]:sl[3]+1,sl[2]:sl[3]+1] = SM[sl[0]:sl[1]+1,sl[0]:sl[1]+1]

    for [sl1, sl2] in combinations(slices,2):
        M[sl1[2]:sl1[3]+1,sl2[2]:sl2[3]+1] = SM[sl1[0]:sl1[1]+1,sl2[0]:sl2[1]+1]
        M[sl2[2]:sl2[3]+1,sl1[2]:sl1[3]+1] = SM[sl2[0]:sl2[1]+1,sl1[0]:sl1[1]+1]
        


def grounded_params(vl_name,branch,p,vl_info):
    grounded = vl_info[vl_name]["grounded"]
    for item in grounded:
        if item["link_branch"]==branch and (item["supportN"]<=p<=item["supportK"] or item["supportN"]>=p>=item["supportK"]):
            return {"type":item["type"],"resistance":item["resistance"]}

def Y_matrix_builder(lst_zy,s_i_end):
    Y = sparse.lil_matrix((s_i_end, s_i_end),dtype=np.complex128)

    for item in lst_zy:
        if item["type"]=='bypass' or item["type"]=="to_ground":
            Y_now = item.get("Ybp", [1/R_bypass]*len(item["points"]))
            for (p1,p2),Y_cur in zip(item["points"],Y_now):
                Y[p1,p1] += Y_cur
                Y[p2,p2] += Y_cur
                Y[p1,p2] -= Y_cur
                Y[p2,p1] -= Y_cur

            if "ground" in item:
               Y[item["ground"],item["ground"]]+= item["Yzy"]

    return Y


def chain_sort(arr,key=lambda x:(x[0],x[1])):
    new_arr = [*arr[:1]]
    work = True
    while work:
        for i in range(len(arr)):
            if key(arr[i])[1]==key(new_arr[0])[0]:
                new_arr.insert(0,arr[i])
                break
            elif key(arr[i])[0]==key(new_arr[-1])[1]:
                new_arr.insert(len(new_arr),arr[i])
                break
        else:
            work = False

    if len(arr)==len(new_arr):
        return new_arr
    else:
        raise Exception("Elements of arr don't form a sequence")

find_brances = lambda arr,n:list(filter(lambda x: n in x, arr))

def ps_branch(arr,n):
    lst = find_brances(arr,n)

    if len(lst)!=1:
        raise Exception

    branch = lst[0]
    return (branch,1 if branch[0]==n else -1)

def other_branches(arr,n):
    return [(i,1 if i[0]==n else -1) for i in find_brances(arr,n)]

def support_range(n,k,r):
    if n>k and r==1:
        return range(n,k-1,-1)
    elif n<k and r==1:
        return range(n,k+1)
    elif n>k and r==-1:
        return range(k,n+1)
    elif n<k and r==-1:
        return range(k,n-1,-1)

def recursion_add_length(ps,l,node,vl_name,ps_name,branch_l,result):          
    for s in branch_l[ps[0]][::ps[1]]:
        dl = s[2]/abs(s[1]-s[0])
        for p in support_range(*s[:2],ps[1]):
            if (ps[0],p) not in result[vl_name][ps_name]:
                result[vl_name][ps_name][(ps[0],p)]=l
                l+=dl

    l-=dl
    new_node = ps[0][1] if ps[0][0]==node else ps[0][0]
    other = other_branches([i for i in branch_l.keys() if i!=ps[0]], new_node)
    

    for i in other:
        if i[0]!=ps[0]:
            recursion_add_length(i,l,new_node,vl_name,ps_name,branch_l,result)


def length_to_ps_builder(vl_info):
    result = {}
    k1 = lambda i: (i["supportN"],i["supportK"],abs(i["lengthK"]-i["lengthN"]))
    k2 = lambda i: (i["supportN"],i["supportK"],abs(i["length"]))

    for vl_name, params in vl_info.items():
        ps_nodes = {}
        branch_l = {}
        for branch, br_dt in params["branches"].items():
            if br_dt["PS"] == "left":
                ps_nodes[br_dt["PS_name_1"]]=branch[0]
            elif br_dt["PS"] == "right":
                ps_nodes[br_dt["PS_name_2"]]=branch[1]
            if br_dt["PS"] == "both":
                ps_nodes[br_dt["PS_name_1"]]=branch[0]
                ps_nodes[br_dt["PS_name_2"]]=branch[1]
            
            sectors = chain_sort([sect for sect in params["sectors"] if sect["link_branch"]==branch],key=lambda x:(x["supportN"],x["supportK"]))
            branch_l[branch] = [k1(i) if i["type"]=="with_okgt" else k2(i) for i in sectors]

        result[vl_name] = {}
        
        for ps_name, node in ps_nodes.items():
            result[vl_name][ps_name] = {}
            l = 0
            ps = ps_branch(list(branch_l.keys()),node)

            recursion_add_length(ps,l,node,vl_name,ps_name,branch_l,result)

    return result   

def Isc_func2(x,a,b,c):
    return c+b/(x+a) 

def Isc_func(x,a,b,c,d,e,f):
    return a+b*x+c*x**2+d*x**3+e*x**4+f*x**5

def I_sc_corector(I_sc, L_sc, borders=False):
    try:
        for i1, (I, L) in enumerate(zip(I_sc, L_sc)):
            if I is not None or L is not None:
                break

        else:
            if borders:
                return I_sc, L_sc, None
            else:
                return I_sc, L_sc
        
        for i2, I, L in zip(range(len(I_sc),0,-1),I_sc[::-1], L_sc[::-1]):
            if I is not None or L is not None:
                break
        else:
            if borders:
                return I_sc, L_sc, None
            else:
                return I_sc, L_sc
        
        I_sc1, L_sc1 = I_sc[i1:i2], L_sc[i1:i2]

        if I_sc1[0]>I_sc1[-1]:
            L_sc2 =  L_sc[:i1] + [(None if i is None else (i-L_sc1[0] if L_sc1[0]<L_sc1[-1] else L_sc1[0]-i)) for i in L_sc1] + L_sc[i2:]
            I_sc2 = I_sc

        else:
            I_sc1.reverse()
            L_sc1.reverse()
            L_sc2 = L_sc[i2:][::-1] + [(None if i is None else (i-L_sc1[0] if L_sc1[0]<L_sc1[-1] else L_sc1[0]-i)) for i in L_sc1] + L_sc[:i1][::-1]
            I_sc2 = I_sc[i2:][::-1] + I_sc1 + I_sc[:i1][::-1]

        
        if borders:
            return I_sc2, L_sc2, (L_sc2[i1],L_sc2[i2-1])
        else:
            return I_sc2, L_sc2
    except Exception as ex:
        print(ex)
        if borders:
            return I_sc, L_sc, None
        else:
            return I_sc, L_sc
    

def Isc_get_maker(rpa_info):  
    result = {} 
    for key, val in rpa_info.items():
        I_sc = val["I_sc"]
        L_sc = val["L_sc"]

        """ if I_sc[0]>I_sc[-1]:
            L_sc = [i-L_sc[0] if L_sc[0]<L_sc[-1] else L_sc[0]-i for i in L_sc] 

        else:
            I_sc.reverse()
            L_sc.reverse()
            L_sc = [i-L_sc[0] if L_sc[0]<L_sc[-1] else L_sc[0]-i for i in L_sc] """

        I_sc, L_sc = I_sc_corector(I_sc, L_sc)
        
        interp = interp1d(L_sc, I_sc)
        try:
            popt, _ = curve_fit(Isc_func, L_sc, I_sc, maxfev=10**6) #, maxfev=10**6
        except Exception:
            result[key] = {"interpFunc":interp,"aproFunc":interp}
        else:
            result[key] = {"interpFunc":interp,"aproFunc":lambda x, p=popt: Isc_func(x,*p)}

    return result


def J_matrix_builder(point_sc,J_make_lst,length_to_ps_lst,rpa_info,Isc_funcs,s_i_end,s_j_end):
    vl_name = point_sc['vl_name']
    branch = point_sc['branch']
    support = point_sc['support']
    ps_lst = point_sc['ps_lst']

    #interpFunc
    #Isc = {i: Isc_funcs[(vl_name,i)]["aproFunc"](length_to_ps_lst[vl_name][i][branch,support]) for i in ps_lst}
    #Isc = {i: Isc_funcs[(vl_name,i)]["interpFunc"](length_to_ps_lst[vl_name][i][branch,support]) for i in ps_lst}

    Isc = {}
    for i in ps_lst:
        if (vl_name,i) not in rpa_info:
            continue
        try:
            Isc[i] = Isc_funcs[(vl_name,i)]["interpFunc"](round(length_to_ps_lst[vl_name][i][branch,support],5))
        except Exception:
            Isc[i] = Isc_funcs[(vl_name,i)]["aproFunc"](length_to_ps_lst[vl_name][i][branch,support])

    T_m = {}

    for ps in ps_lst:
        if (vl_name,ps) not in rpa_info:
            continue
        inf = rpa_info[(vl_name,ps)]
        const_t = inf["Tswitch"] + inf["Tautomation"]
        if len(inf["rpa_I_setting"])==0 and len(inf["rpa_time_setting"])==0:
            I_t = [(Isc[ps],float("inf"))] 
        else:
            Iy =  min(map(lambda x,y=Isc[ps]: (y-x if y>=x else float("inf"),x) , inf["rpa_I_setting"]))
            ind = inf["rpa_I_setting"].index(Iy[1]) if Iy[0]!=float("inf") else None

            if ind is None:
                raise Exception(f"Current of relay settings more than I short {vl_name} {ps}")

            I_t = [(Isc[ps],const_t+inf["rpa_time_setting"][ind])]

        for i in range(inf["arc_times"]):
            I_t.append((0,inf["arc_pause"][i]))
            I_t.append((Isc[ps],const_t+inf["rpa_time_setting"][ind]-inf["arc_setting"][i][ind]))

        T_m[ps] = I_t
    
    Time_line = []
    t = 0
    while True:
        lst_ind = {}
        for ps in T_m:
            dt = 0
            for p in T_m[ps]:
                if dt<=t<dt+p[1]: 
                    lst_ind[ps] = (p[0],dt+p[1]-t)
                    break
                else:
                    dt+=p[1]
            else:
                lst_ind[ps] = (0,float("inf"))

        t_old = t
        
        t += min(lst_ind.items(),key=lambda x: x[1][1])[1][1]

        if t == float("inf"): 
            break
        
        I = {ps:i[0] for ps, i in lst_ind.items()}

        if sum(I.values())!=0:
            Time_line.append((t-t_old,I))
    
    JE = sparse.lil_matrix((s_i_end+s_j_end, len(Time_line)*3),dtype=np.complex128)
    
    t_lst = []
    short = []
    colum = 0
    for dt, Isc in Time_line:
        for i in range(3):
            for ps, I in Isc.items():
                JE[J_make_lst[(vl_name,ps)]['sl1'][0]+i,colum] = I * J_make_lst[(vl_name,ps)]['direction'] *10**3
                JE[J_make_lst[(vl_name,ps)]['sl2'][0]+i,colum] = -1* I * J_make_lst[(vl_name,ps)]['direction'] *10**3
            t_lst.append(dt)
            short.append(i)
            colum += 1

    return JE, t_lst, short

def simple_callback(status,data):
    if status=="Calc equation system":
        print("Составляем систему уравнений")
    elif status=="Calc points":
        current,length,name,branch,support = data
        print(f"Calc point vl: {name}, branch: {branch}, support: {support}, {current} from {length}")



def main_calc(okgt_info, vl_info, ps_info, rpa_info, pz=30, callback=simple_callback, single=None):   
    callback("Calc equation system",None)
    i, j = 0, 0

    okgt_lst = []
    vl_lst = []
    cc_lst = []
    cc_lst_nw = []

    vls_to_okgt = {}
    vls_to_cc = {}
    vl_to_vls = {}

    okgt_to_cc = {}
    to_vls = {}

    okgt_nodes = {}
    vls_nodes = {}
    ps_vls = {}

    okgt_max = {}
    okgt_branch = []
    

    lst_zy = []

    for (n, k), sectors in okgt_info.items():
        okgt_branch.append((n, k))
        if n not in okgt_nodes:
            okgt_nodes[n] = (i,j)

        for ind, sector in enumerate(sectors):
            if sector["type"] == "VL":
                vls_sectors = find_entry(get_vl_sector(sector["name"],vl_info),sector["length"])

                sector_lst = []
                for name, vl_sectors in vls_sectors.items():
                    sector_lst.append([])
                    t = len(sector_lst)-1
                    for vl_sector in vl_sectors:
                        N, K = vl_sector['supportN'], vl_sector['supportK']
                        drct = int((K-N)/abs(K-N))
                        gen = zip(range(N,K,drct),range(N+drct,K+drct,drct))
                        dl = abs(vl_sector['lengthK']-vl_sector['lengthN'])/abs(K-N)

                        for nc, kc in gen:
                            sector_lst[t].append([nc,kc,dl,name,vl_sector['link_branch']])

                for t, vl in enumerate(zip(*sector_lst)):
                    if ind == 0 and t==0 and n in okgt_nodes:
                        i_old = okgt_nodes[n][0]
                        start = (i_old,j,-1)
                    else:
                        start = (i,j,-1)
                    i += 1
                    end = (i,j,1)
                    j += 1

                    for v in vl:
                        vls_to_okgt[(v[3],v[4],v[0],v[1])] = len(okgt_lst)
                    
                    d_row = {
                        "start": start,
                        "end": end,
                        "type":"VL",
                        "length": [i[2] for i in vl],
                        'supportN': [i[0] for i in vl],
                        'supportK': [i[1] for i in vl],
                        "name_vl":[i[3] for i in vl],
                        "link_branch":[i[4] for i in vl],
                        "link_okgt": sector["name"]
                        }
                    okgt_lst.append(d_row)

                
            elif sector["type"] == "single_dielectric":
                if ind == 0 and n in okgt_nodes:
                    i_old = okgt_nodes[n][0]
                    start = (i_old,j,-1)
                else:
                    start = (i,j,-1)
                i += 1
                end = (i,j,1)
                j += 1
                d_row = {"start": start, "end": end, "length":sector["length"], "type":sector["type"]}
                okgt_lst.append(d_row)

                okgt_max[start[1]] = {
                    "conductor": None,
                    "type": sector["type"],
                    "sector_link":sector["name"],
                    "links": [(sector["name"])],
                    "length":sector["length"],
                }

            elif sector["type"] == "single_conductive":
                if sector["way_grounded"] == "not":
                    v = 1 
                elif sector["way_grounded"] in ["left", "right"]:
                    v = sector["point_grounded"]
                elif sector["way_grounded"] == "both":
                    v = sector["point_grounded"]-1
                elif sector["way_grounded"] == "inside":
                    v = sector["point_grounded"]+1
                dl = sector["length"]/v
                
                if sector["countercable"]:
                    Y_cc = 1/R_cc(pz,dl,sector["D_countercable"],sector["H_countercable"])

                else:
                    Y_cc = 0

                grounded = single_okgt_zy(sector,v)
                
                for m in range(v):
                    if ind == 0 and m==0 and n in okgt_nodes:
                        i_old = okgt_nodes[n][0]
                        start = (i_old,j,-1)
                    else:
                        start = (i,j,-1)
                    i += 1
                    end = (i,j,1)
                    j += 1
                    d_row = {
                        "start": start,
                        "end": end,
                        "type": sector["type"],
                        'supportN': m,
                        'supportK': m+1,
                        "link_okgt": sector["name"],
                        "length":dl,
                        "groundwire": sector["groundwire"], 
                        "H_cable":sector["H_cable"],
                        "X_cable":sector["X_cable"],
                        "countercable": sector["countercable"]
                        }
                    okgt_lst.append(d_row)

                    if sector["countercable"]:
                        cc_lst.append({"is_vl":False, "s":start,"e":end,"length":dl,"sector":sector,})
                        okgt_to_cc[len(okgt_lst)-1] = len(cc_lst)-1

                    if grounded[m] is not None:     
                        lst_zy.append({"type":"to_ground","Yzy":type_zy_chose(sector,grounded[m],Y_cc),"points":[],"ground":start[0]})

                    if grounded[v] is not None and m==v-1:      
                        lst_zy.append({"type":"to_ground","Yzy":type_zy_chose(sector,grounded[v],Y_cc),"points":[],"ground":end[0]})

                    okgt_max[start[1]] = {
                        "conductor": sector["groundwire"],
                        "type": sector["type"],
                        "sector_link":sector["name"],
                        "links": [(sector["name"],m,m+1)],
                        "length":dl,
                    }

        if k not in okgt_nodes:
            okgt_nodes[k] = (i,j)


    # Construct graph of transmission lines' links
    
    i += 1

    s_i_vl, s_j_vl = i,j

    for vl_name, vl in vl_info.items():
        nodes_position = {}
        for key, branch in vl["branches"].items():
            if key[0] not in nodes_position:
                nodes_position[key[0]] = (i,j)

            add_Ps = False
            if branch["PS"] in ["both", "left"]:
                add_Ps = True
                nodes_position[key[0]] = (i,j)
            trig_first = False
            order = sector_entry(vl["sectors"],key)
            
            for sector in [vl["sectors"][i] for i in order]:
                
                if key == sector["link_branch"]:
                    
                    N, K = sector["supportN"], sector["supportK"]
                    drct = int((K-N)/abs(K-N))
                    gen = zip(range(abs(K-N)),range(N,K,drct),range(N+drct,K+drct,drct))
                    

                    if sector["type"]=="with_okgt":
                        dl = abs(sector['lengthK']-sector['lengthN'])/abs(K-N)
                    elif sector["type"]=="without_okgt":
                        dl = abs(sector['length'])/abs(K-N)

                    for t, nc, kc in gen:
                        data1,data2,data3,data4,data5 = {},{},{},{},{}

                        keyword = ["name_vl","supportN","supportK","link_branch"]
                        dt = [vl_name, nc, kc, key]
                        for kw,rz in zip(keyword,dt):
                            data1[kw],data2[kw],data3[kw],data4[kw],data5[kw] = rz,rz,rz,rz,rz

                        #"phases", "conductors", "groundwires", "supports"
                        phases = get_vl_sector_info(vl_name,key,(nc, kc,),{"name":"phases"},vl_info)
                        conductors = get_vl_sector_info(vl_name,key,(nc, kc,),{"name":"conductors"},vl_info)
                        supports = get_vl_sector_info(vl_name,key,(nc, kc,),{"name":"supports"},vl_info)
                        groundwires = get_vl_sector_info(vl_name,key,(nc, kc,),{"name":"groundwires"},vl_info)
                        countercables = get_vl_sector_info(vl_name,key,(nc, kc,),{"name":"countercables"},vl_info)
                        

                        keyword = ["link_okgt","type","is_Ps_sector","length","Tsupport","phase","mirrored","conductor"] #"start","end",
                        d1 = [sector.get("link_okgt",None),sector["type"],False,dl,supports["s1"],phases["p1"],phases["mirrored"],conductors["c1"]] #s1,e1,
                        d2 = [sector.get("link_okgt",None),sector["type"],False,dl,supports["s2"],phases["p2"],phases["mirrored"],conductors["c2"]] #s2,e2,
                        d3 = [sector.get("link_okgt",None),sector["type"],False,dl,supports["s3"],phases["p3"],phases["mirrored"],conductors["c3"]] #s3,e3,
                        d4 = [sector.get("link_okgt",None),sector["type"],False,dl,supports["s4"],"T1",phases["mirrored"],[groundwires["gw1"],groundwires["is_okgt"]=="gw1"]] #s4,e4,
                        d5 = [sector.get("link_okgt",None),sector["type"],False,dl,supports["s5"],"T2",phases["mirrored"],[groundwires["gw2"],groundwires["is_okgt"]=="gw2"]] #s5,e5,

                        for kw,rz1,rz2,rz3,rz4,rz5 in zip(keyword,d1,d2,d3,d4,d5):
                            data1[kw],data2[kw],data3[kw],data4[kw],data5[kw] = rz1,rz2,rz3,rz4,rz5

                        if not trig_first and t==0 and key[0] in nodes_position:
                            if not add_Ps:
                                trig_first = True
                                i_old = nodes_position[key[0]][0]
                                s1,s2,s3,s4,s5 = (i_old,j,-1),(i_old+1,j+1,-1),(i_old+2,j+2,-1),(i_old+3,j+3,-1),(i_old+4,j+4,-1)
                                #i+=5
                                e1,e2,e3,e4,e5= (i,j,1),(i+1,j+1,1),(i+2,j+2,1),(i+3,j+3,1),(i+4,j+4,1)
                            else:
                                trig_first = True
                                add_Ps = False
                                i_old = nodes_position[key[0]][0]
                                s1,s2,s3,s4,s5 = (i_old,j,-1),(i_old+1,j+1,-1),(i_old+2,j+2,-1),(i_old+3,j+3,-1),(i_old+4,j+4,-1)
                                i+=5
                                e1,e2,e3,e4,e5= (i,j,1),(i+1,j+1,1),(i+2,j+2,1),(i+3,j+3,1),(i+4,j+4,1)
                                j += 5
                                ps_vls[(vl_name,key,nc,kc)] = (len(vl_lst),branch["PS_name_1"],"left")
                                ps = get_vl_sector_info(vl_name,key,(nc, kc,),{"name":"PSs","side":"PS_name_1"},vl_info)
                                vl_lst.append({"start":s1,"end":e1,**data1,"length":ps["length"],"is_Ps_sector":True})
                                vl_lst.append({"start":s2,"end":e2,**data2,"length":ps["length"],"is_Ps_sector":True})
                                vl_lst.append({"start":s3,"end":e3,**data3,"length":ps["length"],"is_Ps_sector":True})
                                vl_lst.append({"start":s4,"end":e4,**data4,"length":ps["length"],"is_Ps_sector":True})
                                vl_lst.append({"start":s5,"end":e5,**data5,"length":ps["length"],"is_Ps_sector":True})
                                
                                s1,s2,s3,s4,s5 = (i,j,-1),(i+1,j+1,-1),(i+2,j+2,-1),(i+3,j+3,-1),(i+4,j+4,-1)
                                i += 5
                                e1,e2,e3,e4,e5= (i,j,1),(i+1,j+1,1),(i+2,j+2,1),(i+3,j+3,1),(i+4,j+4,1)
                                

                        else:
                            s1,s2,s3,s4,s5 = (i,j,-1),(i+1,j+1,-1),(i+2,j+2,-1),(i+3,j+3,-1),(i+4,j+4,-1)
                            i += 5
                            e1,e2,e3,e4,e5= (i,j,1),(i+1,j+1,1),(i+2,j+2,1),(i+3,j+3,1),(i+4,j+4,1)

                        
                        #e1,e2,e3,e4,e5= (i,j,1),(i+1,j+1,1),(i+2,j+2,1),(i+3,j+3,1),(i+4,j+4,1)
                        j += 5

                        #if sector["type"]=="with_okgt":
                        to_vls[(vl_name,key,nc,kc)] = len(vl_lst)

                        vl_lst.append({"start":s1,"end":e1,**data1})
                        vl_lst.append({"start":s2,"end":e2,**data2})
                        vl_lst.append({"start":s3,"end":e3,**data3})
                        vl_lst.append({"start":s4,"end":e4,**data4})
                        vl_lst.append({"start":s5,"end":e5,**data5})

                        lnk = link_vl_to_vl(vl_info,vl_name,key,nc, kc)
                        if lnk:
                            vl_to_vls[(vl_name,key,nc,kc)] = lnk

                        if countercables is not None:
                            cc_lst.append({"is_vl":True, "vl_name":vl_name,"branch":key,"N":nc,"K":kc, "s":s4,"e":e4,"length":dl,"sector":countercables})


            if branch["PS"] in ["both", "right"]:
                s1,s2,s3,s4,s5 = (i,j,-1),(i+1,j+1,-1),(i+2,j+2,-1),(i+3,j+3,-1),(i+4,j+4,-1)
                i += 5
                e1,e2,e3,e4,e5= (i,j,1),(i+1,j+1,1),(i+2,j+2,1),(i+3,j+3,1),(i+4,j+4,1)
                ps_vls[(vl_name,key,nc,kc)] = (len(vl_lst),branch["PS_name_2"],"right")
                ps = get_vl_sector_info(vl_name,key,(nc, kc,),{"name":"PSs","side":"PS_name_2"},vl_info)
                vl_lst.append({"start":s1,"end":e1,**data1,"length":ps["length"],"is_Ps_sector":True})
                vl_lst.append({"start":s2,"end":e2,**data2,"length":ps["length"],"is_Ps_sector":True})
                vl_lst.append({"start":s3,"end":e3,**data3,"length":ps["length"],"is_Ps_sector":True})
                vl_lst.append({"start":s4,"end":e4,**data4,"length":ps["length"],"is_Ps_sector":True})
                vl_lst.append({"start":s5,"end":e5,**data5,"length":ps["length"],"is_Ps_sector":True})
                
                i+=5
                j+=5
                

            if key[1] not in nodes_position:
                nodes_position[key[1]] = (i,j)

        vls_nodes[vl_name] = nodes_position

    s_i_сс, s_j_сс = i,j

        
    for item in cc_lst:
        if item["is_vl"]:
            vls_to_cc[(item["vl_name"],item["branch"],item["N"],item["K"])] = len(cc_lst_nw)
        d = {"is_vl":item["is_vl"],
            "start":(item["s"][0],j,-1),
            "end":(item["e"][0],j,1),
            "link_s":item["s"],
            "link_e":item["e"],
            "length":item["length"],
            "H_countercable":item["sector"]["H_countercable"],
            "X_countercable": item["sector"]["X_countercable"],
            "D_countercable": item["sector"]["D_countercable"],
            "ground_resistance": item["sector"]["ground_resistance"],
            "connect_to_ps":  item["sector"].get("connect_to_ps",False),
            }
        cc_lst_nw.append(d)
        
        j+=1

    s_i_end, s_j_end = i,j

    

    A = sparse.lil_matrix((s_i_end, s_j_end),dtype=np.float64)
    
    for item in okgt_lst: 
        A[item["start"][:2]] = item["start"][2]
        A[item["end"][:2]] = item["end"][2]

    for item in vl_lst:
        A[item["start"][:2]] = item["start"][2]
        A[item["end"][:2]] = item["end"][2]

    for item in cc_lst_nw:
        A[item["start"][:2]] = item["start"][2]
        A[item["end"][:2]] = item["end"][2]

    

    J_make_lst = {}
    vl_to_ps = {}
    for key, val in ps_vls.items():
        a = {"sl1":(vl_lst[val[0]]["start"][0],vl_lst[val[0]+2]["start"][0]+1),
            "sl2":(vl_lst[val[0]]["end"][0],vl_lst[val[0]+2]["end"][0]+1),
            "direction": 1 if val[2]=="left" else -1,
            }
        
        J_make_lst[(key[0],val[1])] = a

        if key[0] not in vl_to_ps:
            vl_to_ps[key[0]] = []

        vl_to_ps[key[0]].append(val[1])


    Z = sparse.lil_matrix((s_j_end, s_j_end),dtype=np.complex128)
    

    calculated_vl = set()
    trig_maker = lambda lst,t=True: [(not val["conductor"][1] and val["conductor"][0] is not None) and t if i==3 or i==4 else True for i,val in enumerate(lst)]
    key_maker = lambda lst: ["phase" if i!=3 and i!=4 else "groundwire" for i in range(len(lst))]

    trig_maker_ps = lambda lst,t=True: [(val["conductor"][0] is not None) and t if i==3 or i==4 else False for i,val in enumerate(lst)]
    key_maker_ps = lambda lst: ["phase_ps" if i!=3 and i!=4 else "groundwire_ps" for i in range(len(lst))]


    previous_sector = {}

    was_grounded = set()

    okgt_sc = set()
    okgt_sc_lst = []

    
    for i in range(0,len(vl_lst),5):
        key = tuple(vl_lst[i][j] for j in ["name_vl","link_branch","supportN","supportK"])
        if (key,vl_lst[i]['is_Ps_sector']) not in calculated_vl and not vl_lst[i]['is_Ps_sector']:
            isCC = vls_to_cc.get(key)
            isOtherVls = vl_to_vls.get(key,[])
            isOkgt = vls_to_okgt.get(key)

            shift = 0
            trig = trig_maker(vl_lst[i:i+5])
            k = key_maker(vl_lst[i:i+5])
            d = [*vl_lst[i:i+5]]
            calculated_vl.add((key,vl_lst[i]['is_Ps_sector']))
            slices = [(0,4,vl_lst[i]["start"][1],vl_lst[i+4]["start"][1])]
            sl=5

            pos = (3 if vl_lst[i+3]["conductor"][1] else 4) if vl_lst[i]["type"]=="with_okgt" else None

            for j in isOtherVls:
                pt = vl_lst[to_vls[j]:to_vls[j]+5]
                if vl_lst[i]["length"] != pt[0]["length"]:
                    raise Exception(f"Length of common chains in not equal, dl1={vl_lst[i]['length']}, dl2={pt[0]['length']}")
                trig+=trig_maker(pt,t=False)
                k+=key_maker(pt)
                d+=[*pt]
                slices.append((sl,sl+4,pt[0]["start"][1],pt[4]["start"][1]))
                sl+=5
                calculated_vl.add((j,vl_lst[i]['is_Ps_sector']))

            dl = vl_lst[i]["length"]

            if isOkgt is not None:
                k.append("okgt_vl")
                trig.append(True)
                slices.append((sl,sl,okgt_lst[isOkgt]["start"][1],okgt_lst[isOkgt]["start"][1]))
                sl += 1
                for j in range(5,len(d),5):
                    pos_n = 3 if d[j+3]["conductor"][1] else 4
                    if pos != pos_n:
                        raise Exception("OKGT can't be located on diffetent plases of multichain support")
                

                okgt_max[okgt_lst[isOkgt]["start"][1]] = {
                    "conductor": d[pos]["conductor"][0],
                    "type": "VL",
                    "sector_link":d[0]["link_okgt"],
                    "links": [(d[j]["name_vl"],d[j]["link_branch"],d[j]["supportN"],d[j]["supportK"]) for j in range(0,len(d),5)],
                    "length":dl,
                }

                d.append(vl_lst[i+pos])

                
            if isCC is not None:
                #sl+=1
                k.append("cc")
                trig.append(True)  
                slices.append((sl,sl,cc_lst_nw[isCC]["start"][1],cc_lst_nw[isCC]["start"][1]))
                d.append(cc_lst_nw[isCC])
                shift = cc_lst_nw[isCC]["H_countercable"]+0.1


            current_sector = {k:vl_lst[i][k] for k in ["name_vl","link_branch","supportN","supportK","type"]}
            current_sector["link"] = vl_lst[i:i+5]
            current_sector["okgt"] = (pos,isOkgt)
            current_sector["other_vls"] = isOtherVls

            if previous_sector  and current_sector: 
                # connection groundwire to okgt when okgt leaves or joins to vl
                if previous_sector["type"]=="without_okgt" and current_sector["type"]=="with_okgt":
                    if (previous_sector["link_branch"]==current_sector["link_branch"] and\
                        previous_sector["supportK"]==current_sector["supportN"]) or\
                        (previous_sector["link"][0]["end"][0]==current_sector["link"][0]["start"][0]):

                        i_okgt = okgt_lst[current_sector["okgt"][1]]["start"][0]
                        i_vl = previous_sector["link"][current_sector["okgt"][0]]["end"][0]

                        lst_zy.append({'type': 'bypass', 'points': [(i_okgt, i_vl)]})
                        

                elif previous_sector["type"]=="with_okgt" and current_sector["type"]=="without_okgt":
                    if (previous_sector["link_branch"]==current_sector["link_branch"] and\
                        previous_sector["supportK"]==current_sector["supportN"]) or\
                        (previous_sector["link"][0]["end"][0]==current_sector["link"][0]["start"][0]):

                        i_okgt = okgt_lst[previous_sector["okgt"][1]]["end"][0]
                        i_vl = current_sector["link"][previous_sector["okgt"][0]]["start"][0]

                        lst_zy.append({'type': 'bypass', 'points': [(i_okgt, i_vl)]})
                        

                # connection groundwire to other groundwire when vl become multichain or singlechain

                if len(previous_sector["other_vls"]) < len(current_sector["other_vls"]):
                    if (previous_sector["link_branch"]==current_sector["link_branch"] and\
                        previous_sector["supportK"]==current_sector["supportN"]) or\
                        (previous_sector["link"][0]["end"][0]==current_sector["link"][0]["start"][0]):

                        links = set(previous_sector["other_vls"]).symmetric_difference(set(current_sector["other_vls"]))

                        point_lst = []
                        for vl_link in links:
                            pt = vl_lst[to_vls[vl_link]:to_vls[vl_link]+5]

                            i_m1 = current_sector["link"][3]["start"][0]
                            i_m2 = current_sector["link"][4]["start"][0]

                            i_n1 = pt[3]["start"][0]
                            i_n2 = pt[4]["start"][0]

                            point_lst+=[(i_m1,i_n1),(i_m2,i_n2)]
                        if len(links)!=0:
                            lst_zy.append({'type': 'bypass', 'points': point_lst})

                elif len(previous_sector["other_vls"]) > len(current_sector["other_vls"]):
                    if (previous_sector["link_branch"]==current_sector["link_branch"] and\
                        previous_sector["supportK"]==current_sector["supportN"]) or\
                        (previous_sector["link"][0]["end"][0]==current_sector["link"][0]["start"][0]): 

                        links = set(previous_sector["other_vls"]).symmetric_difference(set(current_sector["other_vls"]))

                        point_lst = []
                        for vl_link in links:
                            pt = vl_lst[to_vls[vl_link]:to_vls[vl_link]+5]

                            i_m1 = previous_sector["link"][3]["end"][0]
                            i_m2 = previous_sector["link"][4]["end"][0]

                            i_n1 = pt[3]["end"][0]
                            i_n2 = pt[4]["end"][0]

                            point_lst+=[(i_m1,i_n1),(i_m2,i_n2)]
                        if len(links)!=0:
                            lst_zy.append({'type': 'bypass', 'points': point_lst})
                        
            previous_sector = current_sector

            # 
            st_node = (vl_lst[i]["start"][0])
            end_node = (vl_lst[i]["end"][0])

            if st_node not in was_grounded:
                params = grounded_params(*key[:2],key[2],vl_info)
                if params is not None:
                    was_grounded.add(st_node)
                    i_t1 = vl_lst[i+3]["start"][0]
                    i_t2 = vl_lst[i+4]["start"][0]
                    if isOkgt is not None:
                        i_okgt = okgt_lst[isOkgt]["start"][0]
                        if params["type"]=="first":
                            if pos == 3:
                                lst_zy.append({'type': 'to_ground', 'points': [(i_okgt,i_t1)],"ground":i_t1,"Yzy":1/params["resistance"]})
                            else:
                                lst_zy.append({'type': 'to_ground', 'points': [],"ground":i_t1,"Yzy":1/params["resistance"]})
                        elif params["type"]=="second":
                            if pos == 3:
                                lst_zy.append({'type': 'to_ground', 'points': [],"ground":i_t2,"Yzy":1/params["resistance"]})
                            else:
                                lst_zy.append({'type': 'to_ground', 'points': [(i_okgt,i_t2)],"ground":i_t2,"Yzy":1/params["resistance"]})
                        elif params["type"]=="both":
                            lst_zy.append({'type': 'to_ground', 'points': [(i_okgt,i_t1),(i_t1,i_t2)],"ground":i_t1,"Yzy":1/params["resistance"]})
                    else:
                        if params["type"]=="first":  
                            lst_zy.append({'type': 'to_ground', 'points': [],"ground":i_t1,"Yzy":1/params["resistance"]})  
                        elif params["type"]=="second":
                            lst_zy.append({'type': 'to_ground', 'points': [],"ground":i_t2,"Yzy":1/params["resistance"]})
                        elif params["type"]=="both":
                            lst_zy.append({'type': 'to_ground', 'points': [(i_t1,i_t2)],"ground":i_t1,"Yzy":1/params["resistance"]})

            if end_node not in was_grounded:
                params = grounded_params(*key[:2],key[3],vl_info)
                if params is not None:
                    was_grounded.add(end_node)
                    i_t1 = vl_lst[i+3]["end"][0]
                    i_t2 = vl_lst[i+4]["end"][0]
                    if isOkgt is not None:
                        i_okgt = okgt_lst[isOkgt]["end"][0]
                        if params["type"]=="first":
                            if pos == 3:
                                lst_zy.append({'type': 'to_ground', 'points': [(i_okgt,i_t1)],"ground":i_t1,"Yzy":1/params["resistance"]})
                            else:
                                lst_zy.append({'type': 'to_ground', 'points': [],"ground":i_t1,"Yzy":1/params["resistance"]})
                        elif params["type"]=="second":
                            if pos == 3:
                                lst_zy.append({'type': 'to_ground', 'points': [],"ground":i_t2,"Yzy":1/params["resistance"]})
                            else:
                                lst_zy.append({'type': 'to_ground', 'points': [(i_okgt,i_t2)],"ground":i_t2,"Yzy":1/params["resistance"]})
                        elif params["type"]=="both":
                            lst_zy.append({'type': 'to_ground', 'points': [(i_okgt,i_t1),(i_t1,i_t2)],"ground":i_t1,"Yzy":1/params["resistance"]})
                    else:
                        if params["type"]=="first":  
                            lst_zy.append({'type': 'to_ground', 'points': [],"ground":i_t1,"Yzy":1/params["resistance"]})  
                        elif params["type"]=="second":
                            lst_zy.append({'type': 'to_ground', 'points': [],"ground":i_t2,"Yzy":1/params["resistance"]})
                        elif params["type"]=="both":
                            lst_zy.append({'type': 'to_ground', 'points': [(i_t1,i_t2)],"ground":i_t1,"Yzy":1/params["resistance"]})
            

            if isOkgt is not None:
                for pt in ([vl_lst[i:i+5]]+[vl_lst[to_vls[j]:to_vls[j]+5] for j in isOtherVls]):
                    if (pt[0]["name_vl"],okgt_lst[isOkgt]["start"][0]) not in okgt_sc:
                        sc_point = {
                            "vl_name":pt[0]["name_vl"],
                            "branch":pt[0]["link_branch"],
                            "support":pt[0]["supportN"],
                            "i_okgt":okgt_lst[isOkgt]["start"][0],
                            "i_vl":[pt[j]["start"][0] for j in range(3)],
                            "ps_lst":vl_to_ps[pt[0]["name_vl"]],
                        }
                        okgt_sc_lst.append(sc_point)
                        okgt_sc.add((pt[0]["name_vl"],okgt_lst[isOkgt]["start"][0]))

                    if (pt[0]["name_vl"],okgt_lst[isOkgt]["end"][0]) not in okgt_sc:
                        sc_point = {
                            "vl_name":pt[0]["name_vl"],
                            "branch":pt[0]["link_branch"],
                            "support":pt[0]["supportK"],
                            "i_okgt":okgt_lst[isOkgt]["end"][0],
                            "i_vl":[pt[j]["end"][0] for j in range(3)],
                            "ps_lst":vl_to_ps[pt[0]["name_vl"]],
                        }
                        okgt_sc_lst.append(sc_point)
                        okgt_sc.add((pt[0]["name_vl"],okgt_lst[isOkgt]["end"][0]))


        elif (key,vl_lst[i]['is_Ps_sector']) not in calculated_vl and vl_lst[i]['is_Ps_sector']:
            isOtherVls_ps = [j for j in vl_to_vls.get(key,[]) if j in ps_vls]
            isOkgt = vls_to_okgt.get(key)
            isCC = vls_to_cc.get(key)
            
            shift = 0
            trig = trig_maker_ps(vl_lst[i:i+5])
            k = key_maker_ps(vl_lst[i:i+5])
            d = [*vl_lst[i:i+5]]
            calculated_vl.add((key,vl_lst[i]['is_Ps_sector']))
            slices = [(0,4,vl_lst[i]["start"][1],vl_lst[i+4]["start"][1])]
            sl=5

            i_p1 = vl_lst[i]["start" if ps_vls[key][2]=="left" else "end"][0]
            points_zy = [[i_p1,vl_lst[i+1+j]["start" if ps_vls[key][2]=="left" else "end"][0]] for j in range(4)]

            for j in isOtherVls_ps:
                pt = vl_lst[ps_vls[j][0]:ps_vls[j][0]+5]
                if vl_lst[i]["length"] != pt[0]["length"]:
                    raise Exception(f"Length of common chains in not equal, dl1={vl_lst[i]['length']}, dl2={pt[0]['length']}")
                trig+=trig_maker_ps(pt,t=False)
                k+=key_maker_ps(pt)
                d+=[*pt]
                slices.append((sl,sl+4,pt[0]["start"][1],pt[4]["start"][1]))
                sl+=5
                calculated_vl.add((j,vl_lst[i]['is_Ps_sector']))

                points_zy += [[i_p1,j["start" if ps_vls[key][2]=="left" else "end"][0]] for j in pt]

            
            dl = vl_lst[i]["length"]

            if isOkgt is not None:
                pos = 3 if vl_lst[i+3]["conductor"][1] else 4
                sides = ("end","start") if ps_vls[key][2]=="left" else ("start","end")
                
                node1 = vl_lst[i+pos][sides[0]][0]
                node2 = okgt_lst[isOkgt][sides[1]][0]

                lst_zy.append({"type":"bypass","points":[(node1,node2)]})

            if isCC is not None:
                if cc_lst_nw[isCC]["connect_to_ps"]:
                    R = dl * (4*ro_fe)/(np.pi*cc_lst_nw[isCC]["D_countercable"]**2)*10**3
                    if ps_vls[key][2]=="left":
                        lst_zy.append({"type":"bypass","points":[(i_p1,cc_lst_nw[isCC]["start"][0])], "Ybp":[1/R]})
                    else:
                        lst_zy.append({"type":"bypass","points":[(i_p1,cc_lst_nw[isCC]["end"][0])], "Ybp":[1/R]})
                

            
            lst_zy.append({'type': 'to_ground', 'points': points_zy,"ground":i_p1,"Yzy":1/ps_info[ps_vls[key][1]]["resistance"]})


        else: 
            continue

        X,Y,R,r = [],[],[],[]
        for key, data, tr in zip(k,d,trig):
            Xi,Yi,Ri,ri = init_carson_data(key,data,tr,shift)
            X.append(Xi)
            Y.append(Yi)
            R.append(Ri)
            r.append(ri)

        chek_XY = [(x,y) for tp, x, y in zip(k,X,Y) if tp in ["phase","okgt_vl","cc","groundwire_ps","groundwire"] and x is not None and y is not None]
        
        if len(chek_XY) != len(set(chek_XY)):
            raise Exception("Conductors can't have same coordinates")
            
        submatrix_putter(slices,Carson([X,Y,R,r,trig,pz,dl]), Z)

            
    for i, item in enumerate(okgt_lst):
        if item["type"]=='single_conductive':
            isCC = okgt_to_cc.get(i)

            dl = item["length"]
            trig = [True]
            k = ["okgt_single"]
            d = [item]
            shift = 0
            slices = [(0,0,item["start"][1],item["start"][1])]

            if isCC is not None:
                trig.append(True)
                k.append("cc")
                d.append(cc_lst_nw[isCC])
                shift = cc_lst_nw[isCC]["H_countercable"]+0.1
                slices.append((1,1,cc_lst_nw[isCC]["start"][1],cc_lst_nw[isCC]["start"][1]))

            X,Y,R,r = [],[],[],[]

            for key, data, tr in zip(k,d,trig):
                Xi,Yi,Ri,ri = init_carson_data(key,data,tr,shift)
                X.append(Xi)
                Y.append(Yi)
                R.append(Ri)
                r.append(ri)
            
               
        elif item["type"]=='single_dielectric':
            X,Y,R,r,trig,dl = [None],[None],[R_isol],[None],[False],0
            slices = [(0,0,item["start"][1],item["start"][1])]

        else:
            continue

        chek_XY = [(x,y) for tp, x, y in zip(k,X,Y) if tp in ["phase","okgt_vl","cc","groundwire_ps","groundwire"] and x is not None and y is not None]
        
        if len(chek_XY) != len(set(chek_XY)):
            raise Exception("Conductors can't have same coordinates")
            
        submatrix_putter(slices,Carson([X,Y,R,r,trig,pz,dl]), Z)

    Yadd = Y_matrix_builder(lst_zy,s_i_end)
    Isc_funcs = Isc_get_maker(rpa_info)
    length_to_ps_lst = length_to_ps_builder(vl_info)

    #print(Isc_funcs)
    #for lnnt in [0,6,12,18,24,30,36,42,48,54,60,66,72,74.2]:
        #interpFunc
        #print(lnnt,Isc_funcs[('VL #1', 'PS_1')]['aproFunc'](lnnt))
        #print(lnnt,Isc_funcs[('VL #1', 'PS_1')]['interpFunc'](lnnt))


    # Do matrix operation for create equation system
    dZ = sparse.diags(Z.diagonal(), format='lil')
    W = Z - dZ
    dZ = dZ.tocsc()
    dY=linalg.inv(dZ)

    A = A.tocsc()
    At = A.transpose()
    Yy=(A.dot(dY)).dot(At)

    Yy = Yy.tolil()
    Yy+=Yadd

    KI=A.dot(((W.tocsc()).dot(dY)).transpose())

    YZ=sparse.vstack([sparse.hstack([Yy, KI]), sparse.hstack([At, Z])],format='lil')

    B = np.zeros((s_j_vl,),dtype=np.float64)
    B_f = [np.zeros((s_j_vl,),dtype=np.float64),np.zeros((s_j_vl,),dtype=np.float64),np.zeros((s_j_vl,),dtype=np.float64)]

    Ic = np.zeros((s_j_vl,),dtype=np.float64)
    Ic_f = [np.zeros((s_j_vl,),dtype=np.float64),np.zeros((s_j_vl,),dtype=np.float64),np.zeros((s_j_vl,),dtype=np.float64)]


    for current ,itm in enumerate(okgt_sc_lst):
        
        """ if current != 99:
            continue """

        if single is not None:
            if itm['vl_name']!=single[0] or itm['branch']!=single[1] or itm['support']!=single[2]:
                continue

        JE, t_lst, phase_lst = J_matrix_builder(itm,J_make_lst,length_to_ps_lst,rpa_info,Isc_funcs,s_i_end,s_j_end)
        p1 = itm['i_okgt']
        
        callback("Calc points",(current,len(okgt_sc_lst),itm['vl_name'],itm['branch'],itm['support']))
        
        for i, (t, ph) in enumerate(zip(t_lst, phase_lst)):
            if single is not None:
                if i!=single[3]:
                    continue
            
            p2 = itm['i_vl'][ph]
            #print(ph)

            YZ[p1,p1] += 1/R_bypass
            YZ[p2,p2] += 1/R_bypass
            YZ[p1,p2] -= 1/R_bypass
            YZ[p2,p1] -= 1/R_bypass

            H = linalg.spsolve(YZ.tocsr(),JE[:,i])
            okgt_I = np.abs(H[s_i_end:s_i_end+s_j_vl])
            #B_now = okgt_I**2*t
            B_f[ph] += okgt_I**2*t
            Ic_f[ph] = np.maximum(okgt_I,Ic_f[ph])
            #print(np.shape(B_now))
            #print(s_i_end+s_j_vl-s_i_end)

            #B_now = okgt_I  #**2*t
            #print(JE[:,i],t, ph)
            #print(okgt_I)
            #B = np.maximum(B,B_now)

            YZ[p1,p1] -= 1/R_bypass
            YZ[p2,p2] -= 1/R_bypass
            YZ[p1,p2] += 1/R_bypass
            YZ[p2,p1] += 1/R_bypass

        B = np.maximum(B, np.maximum(B_f[0], np.maximum(B_f[1], B_f[2])))
        Ic = np.maximum(Ic, np.maximum(Ic_f[0], np.maximum(Ic_f[1], Ic_f[2])))
        B_f[0].fill(0)
        B_f[1].fill(0)
        B_f[2].fill(0)
        Ic_f[0].fill(0)
        Ic_f[1].fill(0)
        Ic_f[2].fill(0)

    """ for i in range(np.shape(B)[0]):
        print(round(B[i]/10**6,3))

    print("length",i+1) """

    
    result = {}
    for i, (n,k) in enumerate(okgt_branch):
        dl = 0
        result[(n,k)] = {
            "B":[],
            "Bmax":[],
            "I":[],
            "L":[],
            "type":[],
            "conductor":[],
            "links":[],
            "sectors":[],
            'okgt_types':[],
            'length_to_ps_lst': length_to_ps_lst
        }
        j_s = okgt_nodes[n][1]
        j_e = okgt_nodes[k][1]

        if i==0:
            b = j_s
        else:
            pass 
        
        sector, tp, st = okgt_max[b]["sector_link"], okgt_max[b]["type"], 0
        
        ed = 0
        bad = False
        okgt_t_s = set()
        for j in range(b,j_e):
            Bmax = k_conductors.get(okgt_max[j]['conductor'],{}).get("Bsc",None)
            okgt_t_s.add(okgt_max[j]['conductor'])
            #result[(n,k)]["B"]+= [B[j],B[j]]   #[B[j]/10**6,B[j]/10**6] 
            result[(n,k)]["B"]+= [B[j]/10**6,B[j]/10**6] 
            result[(n,k)]["Bmax"]+=[Bmax,Bmax]
            result[(n,k)]["I"]+=[Ic[j],Ic[j]]

            bad = True if B[j]/10**6>(Bmax if Bmax is not None else -float('inf')) and okgt_max[j]["type"]!="single_dielectric" else bad
            

            result[(n,k)]["conductor"]+=[okgt_max[j]['conductor'],okgt_max[j]['conductor']]
            result[(n,k)]["type"]+=[okgt_max[j]["type"],okgt_max[j]["type"]]
            result[(n,k)]["links"]+=[okgt_max[j]["links"],okgt_max[j]["links"]]
            
            result[(n,k)]["L"].append(dl)
            dl+=okgt_max[j]['length']
            result[(n,k)]["L"].append(dl)
            

            if okgt_max[j]["sector_link"]!=sector:
                result[(n,k)]["sectors"].append((sector,tp,st,ed,bad))
                sector = okgt_max[j]["sector_link"]
                tp = okgt_max[j]["type"]
                st = ed
                bad = False
                result[(n,k)]['okgt_types'].append(okgt_t_s)
                okgt_t_s = set()

            if j == j_e-1:
                result[(n,k)]["sectors"].append((sector,tp,st,ed+2,bad))
                sector = okgt_max[j]["sector_link"]
                tp = okgt_max[j]["type"]
                bad = False
                result[(n,k)]['okgt_types'].append(okgt_t_s)
                okgt_t_s = set()
                
            ed+=2

            
        
        b = j_e

    #print([val['okgt_types'] for val in result.values()])
    #print([val["sectors"] for val in result.values()])
    #print(result)
    return result


if __name__=='__main__':
    from initial_data import okgt_info, vl_info, ps_info, rpa_info
    import matplotlib.pyplot as plt
    
    result = main_calc(okgt_info, vl_info, ps_info, rpa_info)


    for val in result.values():
        fig = plt.figure() 
        ax = fig.add_subplot(111)
        ax.plot(val["L"],val["B"],'r')
        ax.plot(val["L"],val["Bmax"],'b')

    plt.show()


